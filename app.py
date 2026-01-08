import streamlit as st
import traceback
import time
import gc
from itertools import groupby
import pandas as pd
import math
import io
import os
import shutil
import tempfile
import subprocess
import re
import requests
from datetime import timedelta, datetime, date
from copy import copy

# Excel è™•ç†ç›¸é—œåº«
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage

# =========================================================
# 1. é é¢è¨­å®š (Page Config) - å¿…é ˆæ”¾åœ¨æœ€ä¸Šæ–¹
# =========================================================
st.set_page_config(
    layout="wide",
    page_title="Cue Sheet Pro v112.6 (Sales Alias Added)"
)

# =============================================================================
# å°ˆæ¡ˆåç¨±: Cue Sheet Pro (åª’é«”æ’ç¨‹ç”Ÿæˆç³»çµ±)
# =============================================================================

# =========================================================
# 2. Session State åˆå§‹åŒ– (State Initialization)
# =========================================================
DEFAULT_RAGIC_URL = "https://ap15.ragic.com/liuskyo/cue/2" 
DEFAULT_RAGIC_KEY = "L04zZGhrVmtTV3pqN1VLbUpnOFZMa01NTHh3OUw3RUVlb0ovNXUrTXJsaGJhMWpKOUxHanFUODREMmN1dEZvcw==" 

DEFAULT_STATES = {
    "is_supervisor": False,      # ä¸»ç®¡æ¬Šé™é–‹é—œ
    "rad_share": 100,            # å»£æ’­é ç®—ä½”æ¯”
    "fv_share": 0,               # æ–°é®®è¦–é ç®—ä½”æ¯”
    "cf_share": 0,               # å®¶æ¨‚ç¦é ç®—ä½”æ¯”
    "cb_rad": True,              # å•Ÿç”¨å»£æ’­
    "cb_fv": False,              # å•Ÿç”¨æ–°é®®è¦–
    "cb_cf": False,              # å•Ÿç”¨å®¶æ¨‚ç¦
    "ragic_url": DEFAULT_RAGIC_URL,
    "ragic_key": DEFAULT_RAGIC_KEY,
    "ragic_confirm_state": False # ä¸Šå‚³ç¢ºèªè¦–çª—ç‹€æ…‹
}

for key, default_val in DEFAULT_STATES.items():
    if key not in st.session_state:
        st.session_state[key] = default_val

# =========================================================
# 3. å…¨åŸŸå¸¸æ•¸è¨­å®š (Global Constants)
# =========================================================
GSHEET_SHARE_URL = "https://docs.google.com/spreadsheets/d/1bzmG-N8XFsj8m3LUPqA8K70AcIqaK4Qhq1VPWcK0w_s/edit?usp=sharing"
BOLIN_LOGO_URL = "https://docs.google.com/drawings/d/17Uqgp-7LJJj9E4bV7Azo7TwXESPKTTIsmTbf-9tU9eE/export/png"

FONT_MAIN = "å¾®è»Ÿæ­£é»‘é«”"
BS_THIN = 'thin'
BS_MEDIUM = 'medium'
BS_HAIR = 'hair'
FMT_MONEY = '"$"#,##0_);[Red]("$"#,##0)'
FMT_NUMBER = '#,##0'

REGIONS_ORDER = ["åŒ—å€", "æ¡ƒç«¹è‹—", "ä¸­å€", "é›²å˜‰å—", "é«˜å±", "æ±å€"]
DURATIONS = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60]
REGION_DISPLAY_MAP = {
    "åŒ—å€": "åŒ—å€-åŒ—åŒ—åŸº",
    "æ¡ƒç«¹è‹—": "æ¡ƒå€-æ¡ƒç«¹è‹—",
    "ä¸­å€": "ä¸­å€-ä¸­å½°æŠ•",
    "é›²å˜‰å—": "é›²å˜‰å—å€-é›²å˜‰å—",
    "é«˜å±": "é«˜å±å€-é«˜å±",
    "æ±å€": "æ±å€-å®œèŠ±æ±",
    "å…¨çœé‡è²©": "å…¨çœé‡è²©",
    "å…¨çœè¶…å¸‚": "å…¨çœè¶…å¸‚"
}

# =========================================================
# 4. åŸºç¤å·¥å…·å‡½å¼ (Helper Functions)
# =========================================================

def parse_count_to_int(x):
    if x is None: return 0
    if isinstance(x, (int, float)): return int(x)
    s = str(x)
    m = re.findall(r"[\d,]+", s)
    return int(m[0].replace(",", "")) if m else 0

def safe_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()

def html_escape(s):
    if s is None: return ""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&#39;")

def region_display(region):
    return REGION_DISPLAY_MAP.get(region, region)

def get_sec_factor(media_type, seconds, sec_factors):
    """å–å¾—ç§’æ•¸åŠ æˆä¿‚æ•¸ (Factor)ã€‚"""
    factors = sec_factors.get(media_type)
    if not factors:
        if media_type == "æ–°é®®è¦–": factors = sec_factors.get("å…¨å®¶æ–°é®®è¦–")
        elif media_type == "å…¨å®¶å»£æ’­": factors = sec_factors.get("å…¨å®¶å»£æ’­")
    if not factors: return 1.0
    if seconds in factors: return factors[seconds]
    for base in [10, 20, 15, 30]:
        if base in factors: return (seconds / base) * factors[base]
    return 1.0

def calculate_schedule(total_spots, days):
    if days <= 0: return []
    if total_spots % 2 != 0: total_spots += 1
    base, rem = divmod(total_spots // 2, days)
    sch = [base + (1 if i < rem else 0) for i in range(days)]
    return [x * 2 for x in sch]

def get_remarks_text(sign_deadline, billing_month, payment_date):
    d_str = sign_deadline.strftime("%Y/%m/%d (%a)") if sign_deadline else "____/__/__ (__)"
    p_str = payment_date.strftime("%Y/%m/%d") if payment_date else "____/__/__"
    return [
        f"1.è«‹æ–¼ {d_str} 11:30å‰ å›ç°½åŠé€²å–®ï¼Œæ–¹å¯é †åˆ©ä¸Šæª”ã€‚",
        "2.ä»¥ä¸Šç¯€ç›®åç¨±å¦‚æœ‰ç•°å‹•ï¼Œä»¥ä¸Šæª”æ™‚ç¯€ç›®åç¨±ç‚ºä¸»ï¼Œå¦‚é‡é›»å°æ™‚æ®µæ»¿æª”ï¼Œä¸Šæª”æ™‚é–“æŒªå¾Œæˆ–æ›´æ›è‡³åŒç´šæ™‚æ®µã€‚",
        "3.é€šè·¯åº—é‹ªæ•¸èˆ‡é–‹æ©Ÿç‡é–‹æ©Ÿç‡è‡³å°‘ä¸ƒæˆ(ä»¥ä¸Š)ã€‚æ¯æ—¥å› åŠ ç›Ÿæ•¸èª¿æ•´ï¼Œæˆ–é‡åº—èˆ–å¹´åº¦å­£åº¦æ”¹è£ã€è¨­å‚™ç¶­è­·å‡ç´šåŠä¿ä¿®ç­‰ç‹€æ³ï¼Œæœƒæœ‰ä¸€å®šå¹…åº¦å¢æ¸›ã€‚",
        "4.è¨—æ’­æ–¹éœ€æ–¼ä¸Šæª”å‰ 5 å€‹å·¥ä½œå¤©ï¼Œæä¾›å»£å‘Šå¸¶(mp3)ã€å½±ç‰‡/å½±åƒ 1920x1080 (mp4)ã€‚",
        f"5.é›™æ–¹åŒæ„è²»ç”¨è«‹æ¬¾æœˆä»½ : {billing_month}ï¼Œå¦‚æœ‰ä¿®æ­£å¿…è¦ï¼Œå°‡å¦è¡ŒE-Mailå‘ŠçŸ¥ï¼Œä¸¦è¦–ç‚ºæ­£å¼åˆç´„ä¹‹ä¸€éƒ¨åˆ†ã€‚",
        f"6.ä»˜æ¬¾å…Œç¾æ—¥æœŸï¼š{p_str}"
    ]

def format_campaign_details(config):
    details = []
    for media, settings in config.items():
        sec_str = ", ".join([f"{s}ç§’({p}%)" for s, p in settings.get("sec_shares", {}).items()])
        reg_str = "å…¨çœè¯æ’­" if settings.get("is_national") else "/".join(settings.get("regions", []))
        info = f"ã€{media}ã€‘ é ç®—ä½”æ¯”: {settings.get('share')}% | ç§’æ•¸åˆ†é…: {sec_str} | å€åŸŸ: {reg_str}"
        details.append(info)
    return "\n".join(details)

# =========================================================
# Ragic API æ•´åˆ
# =========================================================

def upload_to_ragic(api_url, api_key, data_dict, files_dict=None):
    if not api_url or not api_key:
        return False, "API URL æˆ– API Key æœªè¨­å®š"
    base_url = api_url.split("?")[0]
    headers = {"Authorization": f"Basic {api_key}"}
    payload = dict(data_dict)
    payload["api"] = ""   
    payload["v"] = "3"    
    try:
        resp = requests.post(
            base_url, headers=headers, data=payload, files=files_dict, timeout=120
        )
        try:
            j = resp.json()
        except:
            j = None
        if resp.status_code != 200:
            return False, f"HTTP {resp.status_code}: {resp.text[:200]}"
        if not j:
            return False, f"Ragic å›å‚³é JSON æ ¼å¼: {resp.text[:200]}"
        if j.get("status") == "SUCCESS":
            return True, f"âœ… ä¸Šå‚³æˆåŠŸ! Ragic ID: {j.get('ragicId')}"
        return False, f"âŒ Ragic éŒ¯èª¤ (Code: {j.get('code')}): {j.get('msg')}"
    except Exception as e:
        return False, f"âŒ é€£ç·šç•°å¸¸: {str(e)}"

# =========================================================
# ç³»çµ±å·¥å…·: PDF è½‰æª”èˆ‡è³‡æºè®€å–
# =========================================================

def find_soffice_path():
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice: return soffice
    if os.name == "nt":
        candidates = [r"C:\Program Files\LibreOffice\program\soffice.exe", r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"]
        for p in candidates:
            if os.path.exists(p): return p
    return None

@st.cache_data(show_spinner="æ­£åœ¨ä¸‹è¼‰ Logo...", ttl=3600)
def get_cloud_logo_bytes():
    try:
        response = requests.get(BOLIN_LOGO_URL, timeout=10)
        return response.content if response.status_code == 200 else None
    except: return None

@st.cache_data(show_spinner="æ­£åœ¨ç”Ÿæˆ PDF (LibreOffice)...", ttl=3600)
def xlsx_bytes_to_pdf_bytes(xlsx_bytes: bytes):
    soffice = find_soffice_path()
    if not soffice: return None, "Fail", "ä¼ºæœå™¨æœªå®‰è£ LibreOffice"
    try:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = os.path.join(tmp, "cue.xlsx")
            with open(xlsx_path, "wb") as f: f.write(xlsx_bytes)
            subprocess.run([soffice, "--headless", "--nologo", "--convert-to", "pdf:calc_pdf_Export", "--outdir", tmp, xlsx_path], capture_output=True, timeout=60)
            pdf_path = os.path.join(tmp, "cue.pdf")
            if not os.path.exists(pdf_path):
                for fn in os.listdir(tmp):
                    if fn.endswith(".pdf"): pdf_path = os.path.join(tmp, fn); break
            if os.path.exists(pdf_path):
                with open(pdf_path, "rb") as f: return f.read(), "LibreOffice", ""
            return None, "Fail", "LibreOffice æœªç”¢å‡ºæª”æ¡ˆ"
    except Exception as e: return None, "Fail", str(e)
    finally: gc.collect()

# =========================================================
# HTML é è¦½ç”Ÿæˆå¼•æ“
# =========================================================

def generate_html_preview(rows, days_cnt, start_dt, end_dt, c_name, p_display, format_type, remarks, total_list, grand_total, budget, prod):
    eff_days = days_cnt
    header_cls = "bg-sh-head"
    # === ä¿®æ”¹é»ï¼šæ”¹ç”¨ä¸­æ–‡åˆ¤æ–· ===
    if format_type == "æ±å³": header_cls = "bg-dw-head"
    elif format_type == "é‰‘éœ–": header_cls = "bg-bolin-head"
    # ==========================

    date_th1, date_th2 = "", ""
    curr = start_dt
    weekdays = ["ä¸€", "äºŒ", "ä¸‰", "å››", "äº”", "å…­", "æ—¥"]
    for i in range(eff_days):
        wd = curr.weekday()
        bg = "bg-weekend" if wd >= 5 else ""
        date_th1 += f"<th class='{header_cls} col_day'>{curr.day}</th>"
        date_th2 += f"<th class='{bg} col_day'>{weekdays[wd]}</th>"
        curr += timedelta(days=1)

    cols_def = ["Station", "Location", "Program", "Day-part", "Size", "rate<br>(Net)", "Package-cost<br>(Net)"]
    # === ä¿®æ”¹é»ï¼šæ”¹ç”¨ä¸­æ–‡åˆ¤æ–· ===
    if format_type == "è²æ´»": cols_def = ["é »é“", "æ’­å‡ºåœ°å€", "æ’­å‡ºåº—æ•¸", "æ’­å‡ºæ™‚é–“", "ç§’æ•¸/è¦æ ¼", "å–®åƒ¹", "é‡‘é¡"]
    elif format_type == "é‰‘éœ–": cols_def = ["é »é“", "æ’­å‡ºåœ°å€", "æ’­å‡ºåº—æ•¸", "æ’­å‡ºæ™‚é–“", "è¦æ ¼", "å–®åƒ¹", "é‡‘é¡"]
    # ==========================

    th_fixed = "".join([f"<th rowspan='2' class='{header_cls}'>{c}</th>" for c in cols_def])
    th_total_right = f"<th rowspan='2' class='{header_cls}' style='min-width:50px;'>Total<br>Spots</th>"
      
    unique_media = sorted(list(set([r['media'] for r in rows])))
    order_map = {"å…¨å®¶å»£æ’­": 1, "æ–°é®®è¦–": 2, "å®¶æ¨‚ç¦": 3}
    unique_media.sort(key=lambda x: order_map.get(x, 99))
    medium_str = "/".join(unique_media)
      
    tbody = ""
    rows_sorted = sorted(rows, key=lambda x: ({"å…¨å®¶å»£æ’­":1,"æ–°é®®è¦–":2,"å®¶æ¨‚ç¦":3}.get(x["media"],9), x["seconds"]))
    daily_totals = [0] * eff_days

    for key, group in groupby(rows_sorted, lambda x: (x['media'], x['seconds'], x.get('nat_pkg_display', 0))):
        g_list = list(group)
        g_size = len(g_list)
        is_pkg = g_list[0]['is_pkg_member']
        for i, r in enumerate(g_list):
            tbody += "<tr>"
            rate = f"${r['rate_display']:,}" if isinstance(r['rate_display'], (int, float)) else r['rate_display']
            pkg_val_str = ""
            if is_pkg:
                if i == 0:
                    val = f"${r['nat_pkg_display']:,}"
                    pkg_val_str = f"<td class='right' rowspan='{g_size}'>{val}</td>"
            else:
                val = f"${r['pkg_display']:,}" if isinstance(r['pkg_display'], (int, float)) else r['pkg_display']
                pkg_val_str = f"<td class='right'>{val}</td>"
              
            # === ä¿®æ”¹é»ï¼šæ”¹ç”¨ä¸­æ–‡åˆ¤æ–· ===
            if format_type == "è²æ´»":
                sec_txt = f"{r['seconds']}ç§’"
                tbody += f"<td>{r['media']}</td><td>{r['region']}</td><td>{r.get('program_num','')}</td><td>{r['daypart']}</td><td>{sec_txt}</td><td>{rate}</td>{pkg_val_str}"
            elif format_type == "é‰‘éœ–":
                tbody += f"<td>{r['media']}</td><td>{r['region']}</td><td>{r.get('program_num','')}</td><td>{r['daypart']}</td><td>{r['seconds']}ç§’</td><td>{rate}</td>{pkg_val_str}"
            else:
                tbody += f"<td>{r['media']}</td><td>{r['region']}</td><td>{r.get('program_num','')}</td><td>{r['daypart']}</td><td>{r['seconds']}</td><td>{rate}</td>{pkg_val_str}"
            # ==========================
              
            row_spots_sum = 0
            for d_idx, d in enumerate(r['schedule'][:eff_days]):
                tbody += f"<td>{d}</td>"
                row_spots_sum += d
                if d_idx < len(daily_totals): daily_totals[d_idx] += d
            tbody += f"<td style='font-weight:bold; background-color:#f0f0f0;'>{row_spots_sum}</td></tr>"

    total_row_html = "<tr><td colspan='5' style='text-align:center; font-weight:bold; background-color:#e0e0e0;'>Total</td>"
    total_row_html += f"<td style='text-align:center; font-weight:bold; background-color:#e0e0e0;'>${total_list:,}</td>"
    total_row_html += f"<td style='text-align:center; font-weight:bold; background-color:#e0e0e0;'>${budget:,}</td>"
    grand_total_spots = 0
    for day_sum in daily_totals:
        grand_total_spots += day_sum
        total_row_html += f"<td style='font-weight:bold; background-color:#e0e0e0;'>{day_sum}</td>"
    total_row_html += f"<td style='font-weight:bold; background-color:#d0d0d0; border: 2px solid #000;'>{grand_total_spots}</td></tr>"
    tbody += total_row_html

    remarks_html = "<br>".join([html_escape(x) for x in remarks])
    vat = int(round(budget * 0.05))
    footer_html = f"<div style='margin-top:10px; font-weight:bold; text-align:right;'>è£½ä½œè²»: ${prod:,}<br>5% VAT: ${vat:,}<br>Grand Total: ${grand_total:,}</div>"
      
    css = """
    body { font-family: sans-serif; font-size: 10px; background-color: #ffffff; color: #000000; padding: 5px; }
    table { border-collapse: collapse; width: 100%; background-color: #ffffff; }
    th, td { border: 0.5pt solid #000; padding: 4px; text-align: center; white-space: nowrap; color: #000000; }
    .bg-dw-head { background-color: #4472C4; color: white; }
    .bg-sh-head { background-color: white; color: black; font-weight: bold; border-bottom: 2px solid black; }
    .bg-bolin-head { background-color: #F8CBAD; color: black; }
    .bg-weekend { background-color: #FFFFCC; }
    """
    return f"<html><head><style>{css}</style></head><body><div style='margin-bottom:10px;'><b>å®¢æˆ¶åç¨±ï¼š</b>{html_escape(c_name)} &nbsp; <b>Productï¼š</b>{html_escape(p_display)}<br><b>Periodï¼š</b>{start_dt.strftime('%Y.%m.%d')} - {end_dt.strftime('%Y.%m.%d')} &nbsp; <b>Mediumï¼š</b>{html_escape(medium_str)}</div><div style='overflow-x:auto;'><table><thead><tr>{th_fixed}{date_th1}{th_total_right}</tr><tr>{date_th2}</tr></thead><tbody>{tbody}</tbody></table></div>{footer_html}<div style='margin-top:10px; font-size:11px;'><b>Remarksï¼š</b><br>{remarks_html}</div></body></html>"

# =========================================================
# 5. è³‡æ–™è®€å–èˆ‡é‹ç®— (Data Loading & Calculation)
# =========================================================

@st.cache_data(ttl=300)
def load_config_from_cloud(share_url):
    try:
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", share_url)
        if not match: return None, None, None, None, None, "é€£çµæ ¼å¼éŒ¯èª¤"
        file_id = match.group(1)
        def read_sheet(sheet_name):
            url = f"https://docs.google.com/spreadsheets/d/{file_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
            return pd.read_csv(url)
          
        df_store = read_sheet("Stores")
        df_store.columns = [c.strip() for c in df_store.columns]
        store_counts = dict(zip(df_store['Key'], df_store['Display_Name']))
        store_counts_num = dict(zip(df_store['Key'], df_store['Count']))
          
        df_fact = read_sheet("Factors")
        df_fact.columns = [c.strip() for c in df_fact.columns]
        sec_factors = {}
        for _, row in df_fact.iterrows():
            if row['Media'] not in sec_factors: sec_factors[row['Media']] = {}
            sec_factors[row['Media']][int(row['Seconds'])] = float(row['Factor'])
          
        name_map = {"å…¨å®¶æ–°é®®è¦–": "æ–°é®®è¦–", "å…¨å®¶å»£æ’­": "å…¨å®¶å»£æ’­", "å®¶æ¨‚ç¦": "å®¶æ¨‚ç¦"}
        for k, v in name_map.items():
            if k in sec_factors and v not in sec_factors: sec_factors[v] = sec_factors[k]
          
        df_price = read_sheet("Pricing")
        df_price.columns = [c.strip() for c in df_price.columns]
        pricing_db = {}
        for _, row in df_price.iterrows():
            m, r = row['Media'], row['Region']
            if m == "å®¶æ¨‚ç¦":
                if m not in pricing_db: pricing_db[m] = {}
                pricing_db[m][r] = {
                    "List": int(row['List_Price']), "Net": int(row['Net_Price']),
                    "Std_Spots": int(row['Std_Spots']), "Day_Part": row['Day_Part']
                }
            else:
                if m not in pricing_db: pricing_db[m] = {"Std_Spots": int(row['Std_Spots']), "Day_Part": row['Day_Part']}
                pricing_db[m][r] = [int(row['List_Price']), int(row['Net_Price'])]
        
        # === æ–°å¢ï¼šè®€å– Sales åˆ†é  (çœŸå vs ç¶½è™Ÿ) ===
        df_sales = read_sheet("Sales")
        df_sales.columns = [c.strip() for c in df_sales.columns]
        # å‡è¨­æ¬„ä½ç‚º Name (çœŸå) å’Œ Nickname (Ragicç”¨)ï¼Œå»ºç«‹å°ç…§è¡¨
        # å¦‚æœ Google Sheet æ¬„ä½åä¸åŒï¼Œè«‹ä¿®æ­£é€™è£¡
        if 'Name' in df_sales.columns and 'Nickname' in df_sales.columns:
            sales_map = dict(zip(df_sales['Name'], df_sales['Nickname']))
        else:
            # è‹¥æ¬„ä½æ²’è¨­å°ï¼Œåšå€‹é˜²å‘†ï¼ŒKey=Name, Value=Name (éƒ½ç”¨çœŸå)
            sales_map = {name: name for name in df_sales.iloc[:, 0].tolist()}

        return store_counts, store_counts_num, pricing_db, sec_factors, sales_map, None
    except Exception as e: return None, None, None, None, None, f"è®€å–å¤±æ•—: {str(e)}"

# --- æ–°å¢: é‹ç®—é‚è¼¯é¢æ¿æ¸²æŸ“å‡½å¼ ---
def render_logic_panel(logs):
    """
    ç¹ªè£½é‹ç®—é‚è¼¯é¢æ¿
    """
    if not logs:
        # st.warning("å°šç„¡é‹ç®—ç´€éŒ„")
        return

    st.markdown("### ğŸ§® é‹ç®—é‚è¼¯è©³ç´°é¢æ¿ (é€æ˜åŒ–é‹ç®—)")
      
    # å°‡ log åˆ†çµ„é¡¯ç¤º
    for idx, item in enumerate(logs):
        title = f"#{idx+1} ã€{item['media']}ã€‘ {item['seconds']}ç§’ - {item['region']}"
        with st.expander(title, expanded=False):
            # æ‘˜è¦æ•¸æ“šå€
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("åˆ†é…é ç®— (Budget)", f"${int(item['budget']):,}")
            c2.metric("å–®æª”æˆæœ¬ (Unit Cost)", f"${item['unit_cost_actual']:.2f}")
            c3.metric("ç§’æ•¸ä¿‚æ•¸ (Factor)", f"{item['factor']}")
            c4.metric("æœ€çµ‚æª”æ¬¡ (Spots)", item['spots'])
              
            st.markdown("---")
            # è©³ç´°é‚è¼¯å€
            st.markdown("#### 1. åŸºç¤åƒæ•¸")
            st.text(f"â€¢ åª’é«”èˆ‡å€åŸŸ: {item['media']} ({item['region']})")
            st.text(f"â€¢ å¯¦ä½œåƒ¹ (Net Price): ${item['base_net_price']:,} (ä¾æ“š Pricing è¡¨)")
            st.text(f"â€¢ æ¨™æº–æª”æ¬¡ (Std Spots): {item['std_spots']} æª”")
            st.text(f"â€¢ ç§’æ•¸: {item['seconds']}ç§’ (Factor: {item['factor']})")
              
            st.markdown("#### 2. å–®æª”æˆæœ¬è¨ˆç®—")
            st.latex(r"\text{Unit Cost} = \frac{\text{Net Price}}{\text{Std Spots}} \times \text{Factor}")
            st.code(f"{item['base_net_price']} / {item['std_spots']} * {item['factor']} = {item['unit_cost_actual']:.4f}")
              
            st.markdown("#### 3. æª”æ¬¡è¨ˆç®—èˆ‡æ‡²ç½°åˆ¤å®š")
            st.text(f"â€¢ åˆä¼°æª”æ¬¡ = é ç®— / å–®æª”æˆæœ¬ = {item['budget']:.0f} / {item['unit_cost_actual']:.2f} = {item['spots_init_raw']:.2f}")
              
            if item['is_under_target']:
                st.error(f"âš ï¸ è§¸ç™¼æ‡²ç½°æ©Ÿåˆ¶: åˆä¼°æª”æ¬¡ {math.ceil(item['spots_init_raw'])} < æ¨™æº–æª”æ¬¡ {item['std_spots']}")
                st.markdown("**æ‡²ç½°å…§å®¹ï¼šç¸½æª”æ¬¡é™¤ä»¥ 1.1 (è²»ç”¨ä¸è®Šï¼Œæª”æ¬¡è®Šå°‘)**")
                st.latex(r"\text{Final Spots} = \text{Ceil}\left(\frac{\text{Budget}}{\text{Unit Cost} \times 1.1}\right)")
                st.code(f"{item['budget']:.0f} / ({item['unit_cost_actual']:.2f} * 1.1) = {item['spots_final_raw_penalty']:.2f} -> ç„¡æ¢ä»¶é€²ä½ -> {item['spots']}")
            else:
                st.success(f"âœ… ç¬¦åˆæ¨™æº–: åˆä¼°æª”æ¬¡ {math.ceil(item['spots_init_raw'])} >= æ¨™æº–æª”æ¬¡ {item['std_spots']}")
                st.markdown("**ç„¡æ‡²ç½°**")
                st.latex(r"\text{Final Spots} = \text{Ceil}\left(\frac{\text{Budget}}{\text{Unit Cost}}\right)")
              
            if item.get('note'):
                st.info(f"å‚™è¨»: {item['note']}")


def calculate_plan_data(config, total_budget, days_count, pricing_db, sec_factors, store_counts_num, regions_order):
    """
    æ’ç¨‹é‹ç®—æ ¸å¿ƒå‡½å¼ (å·²å¢åŠ é‚è¼¯è¨˜éŒ„åŠŸèƒ½)
    """
    rows, total_list_accum = [], 0
    logs = [] # åˆå§‹åŒ–æ—¥èªŒåˆ—è¡¨

    for m, cfg in config.items():
        # æ ¹æ“šå„åª’é«”çš„é ç®—ä½”æ¯” (Share) åˆ†é…é ç®—
        m_budget_total = total_budget * (cfg["share"] / 100.0)
          
        for sec, sec_pct in cfg["sec_shares"].items():
            s_budget = m_budget_total * (sec_pct / 100.0)
            if s_budget <= 0: continue
              
            factor = get_sec_factor(m, sec, sec_factors)
              
            # --- é‚è¼¯ A: å»£æ’­èˆ‡æ–°é®®è¦– ---
            if m in ["å…¨å®¶å»£æ’­", "æ–°é®®è¦–"]:
                db = pricing_db[m]
                calc_regs = ["å…¨çœ"] if cfg["is_national"] else cfg["regions"]
                display_regs = regions_order if cfg["is_national"] else cfg["regions"]
                  
                # è¨ˆç®—å–®ä½æˆæœ¬ç¸½å’Œ (å¦‚æœæ˜¯å…¨çœï¼Œå°±æ˜¯å„å€å–®åƒ¹åŠ ç¸½æˆ–ç›´æ¥å–å…¨çœåƒ¹)
                # é€™è£¡ç‚ºäº†èƒ½å¤ é‚„åŸ "å¯¦ä½œåƒ¹"ï¼Œæˆ‘å€‘éœ€è¦çŸ¥é“ç”¨çš„æ˜¯å“ªå€‹ Base Net
                # å› ç‚ºç¨‹å¼é‚è¼¯æ˜¯: unit_net_sum = sum( (Region_Net / Std_Spots) * Factor )
                  
                # ç‚ºäº† Log æ¸…æ™°ï¼Œæˆ‘å€‘åæ¨ "ç¸½å¯¦ä½œåƒ¹ (Net Price Sum)"
                base_net_price_sum = 0
                std_spots_ref = db["Std_Spots"] # 4800 or 5040
                  
                # æ ¸å¿ƒé‹ç®—
                unit_net_sum = 0
                for r in calc_regs:
                      # å–å¾—è©²å€åŸŸ(æˆ–å…¨çœ)çš„å¯¦ä½œåƒ¹
                    reg_net = db[r][1] 
                    base_net_price_sum += reg_net
                    # ç®—å‡ºè©²å€åŸŸçš„å–®æª”æˆæœ¬
                    unit_net_sum += (reg_net / std_spots_ref) * factor
                  
                if unit_net_sum == 0: continue
                  
                # è¨ˆç®—æª”æ¬¡ (Spots)
                spots_init_raw = s_budget / unit_net_sum
                spots_init = math.ceil(spots_init_raw)
                  
                is_under_target = spots_init < std_spots_ref
                calc_penalty = 1.1 if is_under_target else 1.0 
                  
                if cfg["is_national"]:
                    row_display_penalty = 1.0
                    total_display_penalty = 1.1 if is_under_target else 1.0
                else:
                    row_display_penalty = 1.1 if is_under_target else 1.0
                    total_display_penalty = 1.0 
                  
                # æœ€çµ‚æª”æ¬¡è¨ˆç®—
                spots_final_raw = s_budget / (unit_net_sum * calc_penalty)
                spots_final = math.ceil(spots_final_raw)
                  
                if spots_final % 2 != 0: spots_final += 1
                if spots_final == 0: spots_final = 2
                  
                # --- [æ–°å¢] è¨˜éŒ„é‹ç®—é‚è¼¯ ---
                logs.append({
                    "media": m,
                    "region": "å…¨çœè¯æ’­" if cfg["is_national"] else "/".join(cfg["regions"]),
                    "seconds": sec,
                    "budget": s_budget,
                    "base_net_price": base_net_price_sum, # ç¸½å¯¦ä½œåƒ¹
                    "std_spots": std_spots_ref,            # æ¨™æº–æª”æ¬¡
                    "factor": factor,                      # ç§’æ•¸ä¿‚æ•¸
                    "unit_cost_actual": unit_net_sum,      # åŠ æˆå¾Œçš„å–®æª”æˆæœ¬
                    "spots_init_raw": spots_init_raw,
                    "is_under_target": is_under_target,
                    "spots_final_raw_penalty": spots_final_raw,
                    "spots": spots_final,
                    "note": "è‹¥é¸å…¨çœè¯æ’­ï¼Œå¯¦ä½œåƒ¹ç‚ºå…¨çœå®šåƒ¹ï¼›è‹¥é¸å€åŸŸï¼Œå‰‡ç‚ºå„å€å¯¦ä½œåƒ¹åŠ ç¸½ã€‚"
                })
                # -------------------------

                # è¨ˆç®—æ¯æ—¥åˆ†é…
                sch = calculate_schedule(spots_final, days_count)
                  
                # è¨ˆç®—å…¨çœæ‰“åŒ…åƒ¹èˆ‡å–®ä¸€å€åŸŸåƒ¹
                nat_pkg_display = 0
                if cfg["is_national"]:
                    nat_list = db["å…¨çœ"][0]
                    nat_unit_price = int((nat_list / db["Std_Spots"]) * factor * total_display_penalty)
                    nat_pkg_display = nat_unit_price * spots_final
                    total_list_accum += nat_pkg_display
                  
                for i, r in enumerate(display_regs):
                    list_price_region = db[r][0]
                    unit_rate_display = int((list_price_region / db["Std_Spots"]) * factor * row_display_penalty)
                    total_rate_display = unit_rate_display * spots_final
                    row_pkg_display = total_rate_display
                    if not cfg["is_national"]: total_list_accum += row_pkg_display
                      
                    rows.append({
                        "media": m, "region": r, "program_num": store_counts_num.get(f"æ–°é®®è¦–_{r}" if m=="æ–°é®®è¦–" else r, 0),
                        "daypart": db["Day_Part"], "seconds": sec, "spots": spots_final, "schedule": sch,
                        "rate_display": total_rate_display, "pkg_display": row_pkg_display,
                        "is_pkg_member": cfg["is_national"], "nat_pkg_display": nat_pkg_display
                    })
              
            # --- é‚è¼¯ B: å®¶æ¨‚ç¦ ---
            elif m == "å®¶æ¨‚ç¦":
                db = pricing_db["å®¶æ¨‚ç¦"]
                base_std = db["é‡è²©_å…¨çœ"]["Std_Spots"]
                base_net_price = db["é‡è²©_å…¨çœ"]["Net"]
                  
                # æ ¸å¿ƒé‹ç®—
                unit_net = (base_net_price / base_std) * factor
                  
                spots_init_raw = s_budget / unit_net
                spots_init = math.ceil(spots_init_raw)
                  
                is_under_target = spots_init < base_std
                penalty = 1.1 if is_under_target else 1.0
                  
                spots_final_raw = s_budget / (unit_net * penalty)
                spots_final = math.ceil(spots_final_raw)
                  
                if spots_final % 2 != 0: spots_final += 1
                  
                # --- [æ–°å¢] è¨˜éŒ„é‹ç®—é‚è¼¯ ---
                logs.append({
                    "media": m,
                    "region": "å…¨çœé‡è²©+è¶…å¸‚",
                    "seconds": sec,
                    "budget": s_budget,
                    "base_net_price": base_net_price,
                    "std_spots": base_std,
                    "factor": factor,
                    "unit_cost_actual": unit_net,
                    "spots_init_raw": spots_init_raw,
                    "is_under_target": is_under_target,
                    "spots_final_raw_penalty": spots_final_raw,
                    "spots": spots_final,
                    "note": f"è¶…å¸‚æª”æ¬¡æœƒä¾ç…§æ¯”ä¾‹è‡ªå‹•è¨ˆç®— (é‡è²©:{spots_final})"
                })
                # -------------------------

                sch_h = calculate_schedule(spots_final, days_count)
                base_list = db["é‡è²©_å…¨çœ"]["List"]
                unit_rate_h = int((base_list / base_std) * factor * penalty)
                total_rate_h = unit_rate_h * spots_final
                total_list_accum += total_rate_h
                  
                rows.append({
                    "media": m, "region": "å…¨çœé‡è²©", "program_num": store_counts_num["å®¶æ¨‚ç¦_é‡è²©"],
                    "daypart": db["é‡è²©_å…¨çœ"]["Day_Part"], "seconds": sec, "spots": spots_final, "schedule": sch_h,
                    "rate_display": total_rate_h, "pkg_display": total_rate_h, "is_pkg_member": False
                })
                  
                # å®¶æ¨‚ç¦è¶…å¸‚çš„æª”æ¬¡æ˜¯ä¾ç…§é‡è²©æ¯”ä¾‹è¨ˆç®—
                spots_s = int(spots_final * (db["è¶…å¸‚_å…¨çœ"]["Std_Spots"] / base_std))
                sch_s = calculate_schedule(spots_s, days_count)
                rows.append({
                    "media": m, "region": "å…¨çœè¶…å¸‚", "program_num": store_counts_num["å®¶æ¨‚ç¦_è¶…å¸‚"],
                    "daypart": db["è¶…å¸‚_å…¨çœ"]["Day_Part"], "seconds": sec, "spots": spots_s, "schedule": sch_s,
                    "rate_display": "è¨ˆé‡è²©", "pkg_display": "è¨ˆé‡è²©", "is_pkg_member": False
                })
                  
    return rows, total_list_accum, logs

# =========================================================
# 6. Excel æ¸²æŸ“å¼•æ“ (Excel Rendering Engines)
# =========================================================

@st.cache_data(show_spinner="æ­£åœ¨ç”Ÿæˆ Excel å ±è¡¨...", ttl=3600)
def generate_excel_from_scratch(format_type, start_dt, end_dt, client_name, product_name, rows, remarks_list, final_budget_val, prod_cost, sales_person):
      
    # Common Excel Styles
    SIDE_THIN, SIDE_MEDIUM, SIDE_HAIR = Side(style=BS_THIN), Side(style=BS_MEDIUM), Side(style=BS_HAIR)
    SIDE_DOUBLE = Side(style='double')
    BORDER_ALL_THIN = Border(top=SIDE_THIN, bottom=SIDE_THIN, left=SIDE_THIN, right=SIDE_THIN)
    BORDER_ALL_MEDIUM = Border(top=SIDE_MEDIUM, bottom=SIDE_MEDIUM, left=SIDE_MEDIUM, right=SIDE_MEDIUM)
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)
    FONT_STD, FONT_BOLD, FONT_TITLE = Font(name=FONT_MAIN, size=12), Font(name=FONT_MAIN, size=14, bold=True), Font(name=FONT_MAIN, size=48, bold=True)
    FILL_WEEKEND = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
      
    def set_border(cell, top=None, bottom=None, left=None, right=None):
        cur = cell.border
        new_top = Side(style=top) if top else cur.top
        new_bottom = Side(style=bottom) if bottom else cur.bottom
        new_left = Side(style=left) if left else cur.left
        new_right = Side(style=right) if right else cur.right
        cell.border = Border(top=new_top, bottom=new_bottom, left=new_left, right=new_right)

    def draw_outer_border_fast(ws, min_r, max_r, min_c, max_c):
        for c in range(min_c, max_c + 1):
            set_border(ws.cell(min_r, c), top=BS_MEDIUM)
            set_border(ws.cell(max_r, c), bottom=BS_MEDIUM)
        for r in range(min_r, max_r + 1):
            set_border(ws.cell(r, min_c), left=BS_MEDIUM)
            set_border(ws.cell(r, max_c), right=BS_MEDIUM)

    # ---------------------------------------------------------
    # Sub-Engine: Dongwu (æ±å³æ ¼å¼)
    # ---------------------------------------------------------
    def render_dongwu_optimized(ws, start_dt, end_dt, rows, budget, prod):
        eff_days = (end_dt - start_dt).days + 1
        spots_col_idx = 7 + eff_days + 1
        total_cols = spots_col_idx

        COL_WIDTHS = {'A': 19.6, 'B': 22.8, 'C': 14.6, 'D': 20.0, 'E': 13.0, 'F': 19.6, 'G': 17.9}
        ROW_HEIGHTS = {1: 61.0, 2: 29.0, 3: 40.0, 4: 40.0, 5: 40.0, 6: 40.0, 7: 40.0, 8: 40.0}
          
        for k, v in COL_WIDTHS.items(): ws.column_dimensions[k].width = v
        for i in range(eff_days): ws.column_dimensions[get_column_letter(8+i)].width = 8.5
        ws.column_dimensions[get_column_letter(spots_col_idx)].width = 13.0
        for r, h in ROW_HEIGHTS.items(): ws.row_dimensions[r].height = h

        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1"); c = ws['A1']; c.value = "Media Schedule"; c.font = FONT_TITLE; c.alignment = ALIGN_CENTER
        unique_secs = sorted(list(set([r['seconds'] for r in rows])))
        p_str = f"{'ã€'.join([f'{s}ç§’' for s in unique_secs])} {product_name}"
        unique_media = sorted(list(set([r['media'] for r in rows])))
        medium_str = "/".join(unique_media)
          
        infos = [("A3", "å®¢æˆ¶åç¨±ï¼š", client_name), ("A4", "Productï¼š", p_str), ("A5", "Period :", f"{start_dt.strftime('%Y. %m. %d')} - {end_dt.strftime('%Y. %m. %d')}"), ("A6", "Medium :", medium_str)]
        for pos, lbl, val in infos:
            c = ws[pos]; c.value = lbl; c.font = FONT_BOLD; c.alignment = Alignment(vertical='center')
            c2 = ws.cell(c.row, 2); c2.value = val; c2.font = FONT_BOLD; c2.alignment = Alignment(vertical='center')
          
        for c_idx in range(1, total_cols + 1): set_border(ws.cell(3, c_idx), top=BS_MEDIUM)
        ws['H6'] = f"{start_dt.month}æœˆ"; ws['H6'].font = Font(name=FONT_MAIN, size=16, bold=True); ws['H6'].alignment = ALIGN_CENTER
          
        headers = [("A","Station"), ("B","Location"), ("C","Program"), ("D","Day-part"), ("E","Size"), ("F","rate\n(Net)"), ("G","Package-cost\n(Net)")]
        for col, txt in headers:
            col_idx = column_index_from_string(col)
            ws.merge_cells(f"{col}7:{col}8"); c7 = ws.cell(7, col_idx); c7.value = txt; c8 = ws.cell(8, col_idx)
            c7.font = FONT_BOLD; c7.alignment = ALIGN_CENTER; c7.border = BORDER_ALL_THIN; c8.border = BORDER_ALL_THIN
            set_border(c7, top=BS_MEDIUM); set_border(c8, bottom=BS_MEDIUM)

        curr = start_dt
        for i in range(eff_days):
            col_idx = 8 + i
            c_d = ws.cell(7, col_idx); c_w = ws.cell(8, col_idx)
            c_d.value = curr; c_d.number_format = 'm/d'; c_w.value = ["ä¸€","äºŒ","ä¸‰","å››","äº”","å…­","æ—¥"][curr.weekday()]
            if curr.weekday() >= 5: c_w.fill = FILL_WEEKEND
            curr += timedelta(days=1)
            c_d.font = FONT_STD; c_w.font = FONT_STD; c_d.alignment = ALIGN_CENTER; c_w.alignment = ALIGN_CENTER
            c_d.border = BORDER_ALL_THIN; c_w.border = BORDER_ALL_THIN
            set_border(c_d, top=BS_MEDIUM); set_border(c_w, bottom=BS_MEDIUM)

        c_spots_7 = ws.cell(7, spots_col_idx); c_spots_7.value = "æª”æ¬¡"; c_spots_8 = ws.cell(8, spots_col_idx)
        ws.merge_cells(start_row=7, start_column=spots_col_idx, end_row=8, end_column=spots_col_idx)
        c_spots_7.font = FONT_BOLD; c_spots_7.alignment = ALIGN_CENTER; c_spots_7.border = BORDER_ALL_THIN; c_spots_8.border = BORDER_ALL_THIN
        set_border(c_spots_7, top=BS_MEDIUM, left=BS_MEDIUM); set_border(c_spots_8, bottom=BS_MEDIUM, left=BS_MEDIUM)
        set_border(ws['A7'], right=BS_MEDIUM); set_border(ws['A8'], right=BS_MEDIUM)

        curr_row = 9; grouped_data = {"å…¨å®¶å»£æ’­": sorted([r for r in rows if r["media"] == "å…¨å®¶å»£æ’­"], key=lambda x: x["seconds"]), "æ–°é®®è¦–": sorted([r for r in rows if r["media"] == "æ–°é®®è¦–"], key=lambda x: x["seconds"]), "å®¶æ¨‚ç¦": sorted([r for r in rows if r["media"] == "å®¶æ¨‚ç¦"], key=lambda x: x["seconds"])}
        total_rate_sum = 0 

        for m_key, data in grouped_data.items():
            if not data: continue
            start_merge = curr_row
            display_name = f"å…¨å®¶ä¾¿åˆ©å•†åº—\n{m_key if m_key!='å®¶æ¨‚ç¦' else ''}å»£å‘Š"
            if m_key == "å®¶æ¨‚ç¦": display_name = "å®¶æ¨‚ç¦"
            elif m_key == "å…¨å®¶å»£æ’­": display_name = "å…¨å®¶ä¾¿åˆ©å•†åº—\né€šè·¯å»£æ’­å»£å‘Š"
            elif m_key == "æ–°é®®è¦–": display_name = "å…¨å®¶ä¾¿åˆ©å•†åº—\næ–°é®®è¦–å»£å‘Š"

            for idx, r in enumerate(data):
                ws.row_dimensions[curr_row].height = 40
                ws.cell(curr_row, 1, display_name).alignment = ALIGN_CENTER
                ws.cell(curr_row, 2, r["region"]).alignment = ALIGN_CENTER
                ws.cell(curr_row, 3, r.get("program_num", 0)).alignment = ALIGN_CENTER
                ws.cell(curr_row, 4, r["daypart"]).alignment = ALIGN_CENTER
                ws.cell(curr_row, 5, f"{r['seconds']}ç§’").alignment = ALIGN_CENTER
                rate = r['rate_display']; pkg = r['pkg_display']
                if isinstance(rate, (int, float)): total_rate_sum += rate
                if r.get("is_pkg_member"): pkg = r['nat_pkg_display'] if idx == 0 else None
                c_rate = ws.cell(curr_row, 6); c_rate.value = rate; c_rate.number_format = FMT_MONEY; c_rate.alignment = ALIGN_CENTER
                if pkg is not None: c_pkg = ws.cell(curr_row, 7); c_pkg.value = pkg; c_pkg.number_format = FMT_MONEY; c_pkg.alignment = ALIGN_CENTER
                row_sum = 0
                for d_idx in range(eff_days):
                    if d_idx < len(r["schedule"]): val = r["schedule"][d_idx]; row_sum += val; c_s = ws.cell(curr_row, 8+d_idx); c_s.value = val; c_s.number_format = FMT_NUMBER; c_s.alignment = ALIGN_CENTER
                ws.cell(curr_row, spots_col_idx, row_sum).alignment = ALIGN_CENTER
                for c_idx in range(1, total_cols + 1): cell = ws.cell(curr_row, c_idx); cell.font = FONT_STD; cell.border = BORDER_ALL_THIN
                curr_row += 1

            ws.merge_cells(start_row=start_merge, start_column=1, end_row=curr_row-1, end_column=1)
            if data[0].get("is_pkg_member"): ws.merge_cells(start_row=start_merge, start_column=7, end_row=curr_row-1, end_column=7)
            for col in [4, 5]:
                m_start = start_merge
                while m_start < curr_row:
                    m_end = m_start; val = ws.cell(m_start, col).value
                    while m_end + 1 < curr_row and ws.cell(m_end+1, col).value == val: m_end += 1
                    if m_end > m_start: ws.merge_cells(start_row=m_start, start_column=col, end_row=m_end, end_column=col)
                    m_start = m_end + 1
            draw_outer_border_fast(ws, start_merge, curr_row-1, 1, total_cols)
            for r in range(start_merge, curr_row): set_border(ws.cell(r, 1), right=BS_MEDIUM); set_border(ws.cell(r, spots_col_idx), left=BS_MEDIUM)

        ws.row_dimensions[curr_row].height = 30
        c_lbl = ws.cell(curr_row, 5, "Total"); c_lbl.alignment = ALIGN_CENTER; c_lbl.font = FONT_BOLD
        c_rate_sum = ws.cell(curr_row, 6, total_rate_sum); c_rate_sum.number_format = FMT_MONEY; c_rate_sum.alignment = ALIGN_CENTER; c_rate_sum.font = FONT_BOLD
        c_val = ws.cell(curr_row, 7, budget); c_val.number_format = FMT_MONEY; c_val.alignment = ALIGN_CENTER; c_val.font = FONT_BOLD
        total_spots_all = 0
        for d_idx in range(eff_days):
            daily_sum = sum([r['schedule'][d_idx] for r in rows if d_idx < len(r['schedule'])]); total_spots_all += daily_sum
            c = ws.cell(curr_row, 8+d_idx); c.value = daily_sum; c.alignment = ALIGN_CENTER; c.font = FONT_STD; c.number_format = FMT_NUMBER
        ws.cell(curr_row, spots_col_idx, total_spots_all).alignment = ALIGN_CENTER; ws.cell(curr_row, spots_col_idx).font = FONT_STD
        for c_idx in range(1, total_cols + 1): set_border(ws.cell(curr_row, c_idx), top=BS_MEDIUM, bottom=BS_MEDIUM, left=BS_THIN, right=BS_THIN)
        set_border(ws.cell(curr_row, 1), left=BS_MEDIUM, right=BS_MEDIUM); set_border(ws.cell(curr_row, spots_col_idx), left=BS_MEDIUM, right=BS_MEDIUM); curr_row += 1

        vat = int(budget * 0.05); grand_total = budget + vat
        footer_items = [("åª’é«”", budget), ("è£½ä½œ", prod), ("5% VAT", vat), ("Grand Total", grand_total)]
        for label, val in footer_items:
            if label == "åª’é«”": continue 
            ws.row_dimensions[curr_row].height = 30
            c_l = ws.cell(curr_row, 6); c_l.value = label; c_l.alignment = ALIGN_LEFT; c_l.font = FONT_STD
            c_v = ws.cell(curr_row, 7); c_v.value = val; c_v.number_format = FMT_MONEY; c_v.alignment = ALIGN_CENTER; c_v.font = FONT_STD
            set_border(c_l, left=BS_MEDIUM, top=BS_THIN, bottom=BS_THIN, right=BS_THIN)
            set_border(c_v, right=BS_MEDIUM, top=BS_THIN, bottom=BS_THIN, left=BS_THIN)
            if label == "Grand Total":
                for c_idx in range(1, total_cols + 1): set_border(ws.cell(curr_row, c_idx), top=BS_MEDIUM, bottom=BS_MEDIUM)
            curr_row += 1
          
        draw_outer_border_fast(ws, 7, curr_row-1, 1, total_cols); curr_row += 1
        ws.cell(curr_row, 1, "Remarks:").font = Font(name=FONT_MAIN, size=16, bold=True, underline='single')
        for rm in remarks_list:
            curr_row += 1
            is_red = rm.strip().startswith("1.") or rm.strip().startswith("4.")
            c = ws.cell(curr_row, 1); c.value = rm; c.font = Font(name=FONT_MAIN, size=14, color="FF0000" if is_red else "000000")

        curr_row += 2; sig_start = curr_row
        ws.merge_cells(start_row=sig_start, start_column=1, end_row=sig_start, end_column=7); ws.cell(sig_start, 1, "ç”²    æ–¹ï¼šæ±å³å»£å‘Šè‚¡ä»½æœ‰é™å…¬å¸").alignment = ALIGN_LEFT
        ws.merge_cells(start_row=sig_start+1, start_column=1, end_row=sig_start+1, end_column=7); ws.cell(sig_start+1, 1, "çµ±ä¸€ç·¨è™Ÿï¼š20935458").alignment = ALIGN_LEFT
        ws.merge_cells(start_row=sig_start+2, start_column=1, end_row=sig_start+2, end_column=7); ws.cell(sig_start+2, 1, sales_person).alignment = ALIGN_LEFT; ws.cell(sig_start+2, 1).font = FONT_STD
          
        right_start_col = 20 # Column T
        ws.merge_cells(start_row=sig_start, start_column=right_start_col, end_row=sig_start, end_column=right_start_col+7); ws.cell(sig_start, right_start_col, f"ä¹™    æ–¹ï¼š{client_name}").alignment = ALIGN_LEFT
        ws.merge_cells(start_row=sig_start+1, start_column=right_start_col, end_row=sig_start+1, end_column=right_start_col+7); ws.cell(sig_start+1, right_start_col, "çµ±ä¸€ç·¨è™Ÿï¼š").alignment = ALIGN_LEFT
        ws.merge_cells(start_row=sig_start+2, start_column=right_start_col, end_row=sig_start+2, end_column=right_start_col+7); ws.cell(sig_start+2, right_start_col, "å®¢æˆ¶ç°½ç« ï¼š").alignment = ALIGN_LEFT
        for c_idx in range(1, total_cols + 1): set_border(ws.cell(sig_start, c_idx), top=BS_THIN)
        return curr_row + 3

    # ---------------------------------------------------------
    # Sub-Engine: Shenghuo (è²æ´»æ•¸ä½æ ¼å¼)
    # ---------------------------------------------------------
    def render_shenghuo_optimized(ws, start_dt, end_dt, rows, budget, prod):
        SIDE_DOUBLE = Side(style='double')
        eff_days = (end_dt - start_dt).days + 1
        end_c_start = 6 + eff_days
        total_cols = end_c_start + 2

        ws.column_dimensions['A'].width = 22.5; ws.column_dimensions['B'].width = 24.5; ws.column_dimensions['C'].width = 13.8; ws.column_dimensions['D'].width = 19.4; ws.column_dimensions['E'].width = 15.0
        for i in range(eff_days): ws.column_dimensions[get_column_letter(6 + i)].width = 8.1 
        ws.column_dimensions[get_column_letter(end_c_start)].width = 9.5; ws.column_dimensions[get_column_letter(end_c_start+1)].width = 58.0; ws.column_dimensions[get_column_letter(end_c_start+2)].width = 20.0 
        ROW_H_MAP = {1:30, 2:30, 3:46, 4:46, 5:40, 6:40, 7:35, 8:35}; 
        for r, h in ROW_H_MAP.items(): ws.row_dimensions[r].height = h
          
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1"); c1 = ws['A1']; c1.value = "è²æ´»æ•¸ä½-åª’é«”è¨ˆåŠƒæ’ç¨‹è¡¨"; c1.font = Font(name=FONT_MAIN, size=24, bold=True); c1.alignment = ALIGN_CENTER
        ws.merge_cells(f"A2:{get_column_letter(total_cols)}2"); c2 = ws['A2']; c2.value = "Media Schedule"; c2.font = Font(name=FONT_MAIN, size=18, bold=True); c2.alignment = ALIGN_CENTER
        FONT_16 = Font(name=FONT_MAIN, size=16); ws.merge_cells(f"A3:{get_column_letter(total_cols)}3"); ws['A3'].value = "è²æ´»æ•¸ä½ç§‘æŠ€è‚¡ä»½æœ‰é™å…¬å¸ çµ±ç·¨ 28710100"; ws['A3'].font = FONT_16; ws['A3'].alignment = ALIGN_LEFT
        ws.merge_cells(f"A4:{get_column_letter(total_cols)}4"); ws['A4'].value = sales_person; ws['A4'].font = FONT_16; ws['A4'].alignment = ALIGN_LEFT
          
        unique_secs = sorted(list(set([r['seconds'] for r in rows]))); sec_str = " ".join([f"{s}ç§’å»£å‘Š" for s in unique_secs]); period_str = f"åŸ·è¡ŒæœŸé–“ï¼š{start_dt.strftime('%Y.%m.%d')} - {end_dt.strftime('%Y.%m.%d')}"
        FONT_14 = Font(name=FONT_MAIN, size=14); c5a = ws['A5']; c5a.value = "å®¢æˆ¶åç¨±ï¼š"; c5a.font = FONT_14; c5a.alignment = ALIGN_LEFT
        ws.merge_cells("B5:E5"); c5b = ws['B5']; c5b.value = client_name; c5b.font = FONT_14; c5b.alignment = ALIGN_LEFT
        ws.merge_cells(f"F5:{get_column_letter(end_c_start)}5"); c5f = ws['F5']; c5f.value = f"å»£å‘Šè¦æ ¼ï¼š{sec_str}"; c5f.font = FONT_14; c5f.alignment = ALIGN_LEFT
        ws.merge_cells(f"{get_column_letter(end_c_start+1)}5:{get_column_letter(total_cols)}5"); c5_r = ws[f"{get_column_letter(end_c_start+1)}5"]; c5_r.value = period_str; c5_r.font = FONT_14; c5_r.alignment = ALIGN_LEFT 
        draw_outer_border_fast(ws, 5, 5, 1, total_cols)

        c6a = ws['A6']; c6a.value = "å»£å‘Šåç¨±ï¼š"; c6a.font = FONT_14; c6a.alignment = ALIGN_LEFT; ws.merge_cells("B6:E6"); c6b = ws['B6']; c6b.value = product_name; c6b.font = FONT_14; c6b.alignment = ALIGN_LEFT
        month_groups = []
        for i in range(eff_days):
            d = start_dt + timedelta(days=i); m_key = (d.year, d.month)
            if not month_groups or month_groups[-1][0] != m_key: month_groups.append([m_key, i, i]) 
            else: month_groups[-1][2] = i
        for m_key, s_idx, e_idx in month_groups:
            start_col = 6 + s_idx; end_col = 6 + e_idx
            ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=end_col); c = ws.cell(6, start_col); c.value = f"{m_key[1]}æœˆ"; c.font = FONT_BOLD; c.alignment = ALIGN_LEFT; c.border = BORDER_ALL_MEDIUM
        for c_idx in range(1, total_cols + 1):
            c = ws.cell(6, c_idx); t, b, l, r = BS_MEDIUM, BS_MEDIUM, None, None
            if c_idx == 1: l = BS_MEDIUM 
            if c_idx == total_cols: r = BS_MEDIUM 
            if c_idx == 6: l = None 
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l) if l else None, right=Side(style=r) if r else None)
        draw_outer_border_fast(ws, 6, 6, 1, 5); ws.cell(6, 5).border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_MEDIUM), right=Side(style=None))
          
        header_start_row = 7; headers = ["é »é“", "æ’­å‡ºåœ°å€", "æ’­å‡ºåº—æ•¸", "æ’­å‡ºæ™‚é–“", "ç§’æ•¸\nè¦æ ¼"]
        for i, h in enumerate(headers):
            c_idx = i + 1; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == 1: l = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r))

        curr = start_dt
        for i in range(eff_days):
            col_idx = 6 + i; c7 = ws.cell(header_start_row, col_idx); c7.value = curr.day; c7.font = FONT_BOLD; c7.alignment = ALIGN_CENTER; c7.border = BORDER_ALL_MEDIUM
            c7.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            c8 = ws.cell(header_start_row+1, col_idx); c8.value = ["æ—¥","ä¸€","äºŒ","ä¸‰","å››","äº”","å…­"][(curr.weekday()+1)%7]; c8.font = FONT_BOLD; c8.alignment = ALIGN_CENTER
            style_left = BS_MEDIUM if col_idx == 6 else BS_THIN
            c8.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=style_left), right=Side(style=BS_THIN)); 
            if curr.weekday() >= 5: c8.fill = FILL_WEEKEND
            curr += timedelta(days=1)

        end_headers = ["æª”æ¬¡", "å®šåƒ¹", "å°ˆæ¡ˆåƒ¹"]; 
        for i, h in enumerate(end_headers):
            c_idx = end_c_start + i; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == total_cols: r = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r))

        date_start_col = 6
        for c_idx in range(date_start_col, total_cols + 1):
            c7 = ws.cell(header_start_row, c_idx); c7.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            if c_idx == date_start_col: set_border(c7, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c7, right=BS_MEDIUM)
            c8 = ws.cell(header_start_row+1, c_idx); c8.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            if c_idx == date_start_col: set_border(c8, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c8, right=BS_MEDIUM)

        curr_row = header_start_row + 2; grouped_data = {"å…¨å®¶å»£æ’­": sorted([r for r in rows if r["media"]=="å…¨å®¶å»£æ’­"], key=lambda x:x['seconds']), "æ–°é®®è¦–": sorted([r for r in rows if r["media"]=="æ–°é®®è¦–"], key=lambda x:x['seconds']), "å®¶æ¨‚ç¦": sorted([r for r in rows if r["media"]=="å®¶æ¨‚ç¦"], key=lambda x:x['seconds'])}
        total_store_count = 0; total_list_sum = 0

        for m_key, data in grouped_data.items():
            if not data: continue
            start_merge = curr_row; d_name = f"å…¨å®¶ä¾¿åˆ©å•†åº—\n{m_key}å»£å‘Š" if m_key != "å®¶æ¨‚ç¦" else "å®¶æ¨‚ç¦"
            for idx, r in enumerate(data):
                ws.row_dimensions[curr_row].height = 40; ws.cell(curr_row, 1, d_name).alignment = ALIGN_CENTER; ws.cell(curr_row, 2, r['region']).alignment = ALIGN_CENTER
                p_num = int(r.get('program_num', 0)); total_store_count += p_num; suffix = "é¢" if m_key == "æ–°é®®è¦–" else "åº—"; ws.cell(curr_row, 3, f"{p_num:,}{suffix}").alignment = ALIGN_CENTER
                ws.cell(curr_row, 4, r['daypart']).alignment = ALIGN_CENTER
                sec = r['seconds']; sec_txt = f"{sec}ç§’\nå½±ç‰‡/å½±åƒ 1920x1080 (mp4)" if m_key == "æ–°é®®è¦–" else f"{sec}ç§’å»£å‘Š"; c_spec = ws.cell(curr_row, 5, sec_txt); c_spec.alignment = ALIGN_CENTER; c_spec.font = Font(name=FONT_MAIN, size=10)
                row_sum = 0
                for d_idx in range(eff_days):
                    if d_idx < len(r['schedule']): val = r['schedule'][d_idx]; row_sum += val; c = ws.cell(curr_row, 6+d_idx); c.value = val; c.alignment = ALIGN_CENTER; c.font = FONT_STD; c.border = BORDER_ALL_THIN
                ws.cell(curr_row, end_c_start, row_sum).alignment = ALIGN_CENTER
                rate_val = r['rate_display']; 
                if isinstance(rate_val, (int, float)): total_list_sum += rate_val
                ws.cell(curr_row, end_c_start+1, rate_val).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER 
                pkg = r['pkg_display']; 
                if r.get('is_pkg_member'): pkg = r['nat_pkg_display'] if idx == 0 else None
                ws.cell(curr_row, end_c_start+2, pkg).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+2).alignment = ALIGN_CENTER
                for c_idx in range(1, total_cols + 1): c = ws.cell(curr_row, c_idx); c.font = FONT_STD; c.border = BORDER_ALL_THIN
                set_border(ws.cell(curr_row, 5), right=BS_MEDIUM); curr_row += 1
            ws.merge_cells(start_row=start_merge, start_column=1, end_row=curr_row-1, end_column=1)
            if data[0].get('is_pkg_member'): ws.merge_cells(start_row=start_merge, start_column=end_c_start+2, end_row=curr_row-1, end_column=end_c_start+2)
            draw_outer_border_fast(ws, start_merge, curr_row-1, 1, total_cols)

        ws.row_dimensions[curr_row].height = 40; ws.cell(curr_row, 3, total_store_count).number_format = FMT_NUMBER; ws.cell(curr_row, 3).alignment = ALIGN_CENTER; ws.cell(curr_row, 3).font = FONT_BOLD
        ws.cell(curr_row, 5, "Total").alignment = ALIGN_CENTER; ws.cell(curr_row, 5).font = FONT_BOLD
        for d_idx in range(eff_days): daily_sum = sum([r['schedule'][d_idx] for r in rows if d_idx < len(r['schedule'])]); c = ws.cell(curr_row, 6+d_idx); c.value = daily_sum; c.alignment = ALIGN_CENTER; c.font = FONT_BOLD
        ws.cell(curr_row, end_c_start, sum([sum(r['schedule']) for r in rows])).alignment = ALIGN_CENTER; ws.cell(curr_row, end_c_start).font = FONT_BOLD
        ws.cell(curr_row, end_c_start+1, total_list_sum).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).font = FONT_BOLD; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER
        ws.cell(curr_row, end_c_start+2, budget).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+2).font = FONT_BOLD; ws.cell(curr_row, end_c_start+2).alignment = ALIGN_CENTER
        for c_idx in range(1, total_cols+1): ws.cell(curr_row, c_idx).border = BORDER_ALL_THIN
        draw_outer_border_fast(ws, curr_row, curr_row, 1, total_cols)
        for c_idx in range(1, total_cols+1): set_border(ws.cell(curr_row, c_idx), bottom=BS_MEDIUM)
        set_border(ws.cell(curr_row, 5), right=BS_MEDIUM); curr_row += 1

        vat = int(budget * 0.05); grand_total = budget + vat
        footer_stack = [("è£½ä½œ", prod), ("5% VAT", vat), ("Grand Total", grand_total)]
        for lbl, val in footer_stack:
            ws.row_dimensions[curr_row].height = 30; c_l = ws.cell(curr_row, end_c_start+1); c_l.value = lbl; c_l.alignment = ALIGN_RIGHT; c_l.font = FONT_STD
            c_v = ws.cell(curr_row, end_c_start+2); c_v.value = val; c_v.number_format = FMT_MONEY; c_v.alignment = ALIGN_CENTER; c_v.font = FONT_BOLD 
            t, b, l, r = BS_THIN, BS_THIN, BS_MEDIUM, BS_THIN; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_l.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            t, b, l, r = BS_THIN, BS_THIN, BS_THIN, BS_MEDIUM; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_v.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            if lbl == "Grand Total":
                for c_idx in range(1, total_cols + 1): set_border(ws.cell(curr_row, c_idx), bottom=BS_MEDIUM)
            curr_row += 1
          
        curr_row += 1; start_footer = curr_row; r_col_start = 6 
        ws.row_dimensions[start_footer].height = 25; ws.cell(start_footer, r_col_start).value = "Remarksï¼š"
        ws.cell(start_footer, r_col_start).font = Font(name=FONT_MAIN, size=16, bold=True)
        r_row = start_footer
        for rm in remarks_list:
            r_row += 1; ws.row_dimensions[r_row].height = 25; is_red = rm.strip().startswith("1.") or rm.strip().startswith("4."); is_blue = rm.strip().startswith("6."); color = "000000"
            if is_red: color = "FF0000"
            if is_blue: color = "0000FF"
            c = ws.cell(r_row, r_col_start); c.value = rm; c.font = Font(name=FONT_MAIN, size=16, color=color)

        sig_col_start = 1
        ws.cell(start_footer, sig_col_start).value = "ä¹™        æ–¹ï¼š"; ws.cell(start_footer, sig_col_start).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+1, sig_col_start+1).value = client_name; ws.cell(start_footer+1, sig_col_start+1).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+2, sig_col_start).value = "çµ±ä¸€ç·¨è™Ÿï¼š"; ws.cell(start_footer+2, sig_col_start).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+2, sig_col_start+2).value = ""; ws.cell(start_footer+2, sig_col_start+2).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+3, sig_col_start).value = "å®¢æˆ¶ç°½ç« ï¼š"; ws.cell(start_footer+3, sig_col_start).font = Font(name=FONT_MAIN, size=16)

        target_border_row = r_row + 2
        for c_idx in range(1, total_cols + 1): ws.cell(target_border_row, c_idx).border = Border(bottom=SIDE_DOUBLE)
        return target_border_row

    # ---------------------------------------------------------
    # Sub-Engine: Bolin (é‰‘éœ–æ ¼å¼)
    # ---------------------------------------------------------
    def render_bolin_optimized(ws, start_dt, end_dt, rows, budget, prod):
        SIDE_DOUBLE = Side(style='double')
        logo_bytes = get_cloud_logo_bytes()
        eff_days = (end_dt - start_dt).days + 1; end_c_start = 6 + eff_days; total_cols = end_c_start + 2
        ws.column_dimensions['A'].width = 21.0; ws.column_dimensions['B'].width = 21.0; ws.column_dimensions['C'].width = 13.8; ws.column_dimensions['D'].width = 19.4; ws.column_dimensions['E'].width = 15.0
        for i in range(eff_days): ws.column_dimensions[get_column_letter(6 + i)].width = 8.1
        ws.column_dimensions[get_column_letter(end_c_start)].width = 9.5; ws.column_dimensions[get_column_letter(end_c_start+1)].width = 36.0; ws.column_dimensions[get_column_letter(end_c_start+2)].width = 20.0
        ROW_H_MAP = {1:70, 2:33.5, 3:33.5, 4:46, 5:40, 6:35, 7:35}
        for r, h in ROW_H_MAP.items(): ws.row_dimensions[r].height = h
          
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1"); c1 = ws['A1']; c1.value = "é‰‘éœ–è¡Œå‹•è¡ŒéŠ·-åª’é«”è¨ˆåŠƒæ’ç¨‹è¡¨ Mobi Media Schedule"; c1.font = Font(name=FONT_MAIN, size=28, bold=True); c1.alignment = ALIGN_LEFT 
        if logo_bytes:
            try: img = OpenpyxlImage(io.BytesIO(logo_bytes)); scale = 125 / img.height; img.height = 125; img.width = int(img.width * scale); col_letter = get_column_letter(total_cols - 1); img.anchor = f"{col_letter}1"; ws.add_image(img)
            except Exception: pass

        c2a = ws['A2']; c2a.value = "TOï¼š"; c2a.font = Font(name=FONT_MAIN, size=20, bold=True, color="FF0000"); c2a.alignment = ALIGN_LEFT
        ws.merge_cells(f"B2:{get_column_letter(total_cols)}2"); c2b = ws['B2']; c2b.value = client_name; c2b.font = Font(name=FONT_MAIN, size=20, bold=True, color="FF0000"); c2b.alignment = ALIGN_LEFT
        c3a = ws['A3']; c3a.value = "FROMï¼š"; c3a.font = Font(name=FONT_MAIN, size=20, bold=True); c3a.alignment = ALIGN_LEFT
        ws.merge_cells(f"B3:{get_column_letter(total_cols)}3"); c3b = ws['B3']; c3b.value = f"é‰‘éœ–è¡Œå‹•è¡ŒéŠ· {sales_person}"; c3b.font = Font(name=FONT_MAIN, size=20, bold=True); c3b.alignment = ALIGN_LEFT

        unique_secs = sorted(list(set([r['seconds'] for r in rows]))); sec_str = " ".join([f"{s}ç§’å»£å‘Š" for s in unique_secs]); period_str = f"åŸ·è¡ŒæœŸé–“ï¼š{start_dt.strftime('%Y.%m.%d')} - {end_dt.strftime('%Y.%m.%d')}"
        c4a = ws['A4']; c4a.value = "å®¢æˆ¶åç¨±ï¼š"; c4a.font = Font(name=FONT_MAIN, size=14, bold=True); c4a.alignment = ALIGN_LEFT
        ws.merge_cells("B4:E4"); c4b = ws['B4']; c4b.value = client_name; c4b.font = Font(name=FONT_MAIN, size=14, bold=True); c4b.alignment = ALIGN_LEFT
        spec_merge_start = "F4"; spec_merge_end = f"{get_column_letter(end_c_start)}4"; ws.merge_cells(f"{spec_merge_start}:{spec_merge_end}"); c4f = ws['F4']; c4f.value = f"å»£å‘Šè¦æ ¼ï¼š{sec_str}"; c4f.font = Font(name=FONT_MAIN, size=14, bold=True); c4f.alignment = ALIGN_LEFT
        ws.merge_cells(f"{get_column_letter(end_c_start+1)}4:{get_column_letter(total_cols)}4"); c4_r = ws[f"{get_column_letter(end_c_start+1)}4"]; c4_r.value = period_str; c4_r.font = Font(name=FONT_MAIN, size=14, bold=True); c4_r.alignment = ALIGN_LEFT
        draw_outer_border_fast(ws, 4, 4, 1, total_cols)

        c5a = ws['A5']; c5a.value = "å»£å‘Šåç¨±ï¼š"; c5a.font = Font(name=FONT_MAIN, size=14, bold=True); c5a.alignment = ALIGN_LEFT
        ws.merge_cells("B5:E5"); c5b = ws['B5']; c5b.value = product_name; c5b.font = Font(name=FONT_MAIN, size=14, bold=True); c5b.alignment = ALIGN_LEFT
        month_groups = []
        for i in range(eff_days):
            d = start_dt + timedelta(days=i); m_key = (d.year, d.month)
            if not month_groups or month_groups[-1][0] != m_key: month_groups.append([m_key, i, i]) 
            else: month_groups[-1][2] = i
        for m_key, s_idx, e_idx in month_groups:
            start_col = 6 + s_idx; end_col = 6 + e_idx; ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col); c = ws.cell(5, start_col); c.value = f"{m_key[1]}æœˆ"; c.font = FONT_BOLD; c.alignment = ALIGN_LEFT 
        for c_idx in range(1, total_cols + 1):
            c = ws.cell(5, c_idx); t, b, l, r = BS_MEDIUM, BS_MEDIUM, None, None
            if c_idx == 1: l = BS_MEDIUM 
            if c_idx == total_cols: r = BS_MEDIUM 
            if c_idx == 6: l = None 
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l) if l else None, right=Side(style=r) if r else None)
        draw_outer_border_fast(ws, 5, 5, 1, 5); ws.cell(5, 5).border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_MEDIUM), right=Side(style=None))

        header_start_row = 6; headers = ["é »é“", "æ’­å‡ºåœ°å€", "æ’­å‡ºåº—æ•¸", "æ’­å‡ºæ™‚é–“", "ç§’æ•¸\nè¦æ ¼"]
        for i, h in enumerate(headers):
            c_idx = i + 1; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == 1: l = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r))

        curr = start_dt
        for i in range(eff_days):
            col_idx = 6 + i; c6 = ws.cell(header_start_row, col_idx); c6.value = curr.day; c6.font = FONT_BOLD; c6.alignment = ALIGN_CENTER; c6.border = BORDER_ALL_MEDIUM; c6.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            c7 = ws.cell(header_start_row+1, col_idx); c7.value = ["æ—¥","ä¸€","äºŒ","ä¸‰","å››","äº”","å…­"][(curr.weekday()+1)%7]; c7.font = FONT_BOLD; c7.alignment = ALIGN_CENTER; style_left = BS_MEDIUM if col_idx == 6 else BS_THIN; c7.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=style_left), right=Side(style=BS_THIN))
            if curr.weekday() >= 5: c7.fill = FILL_WEEKEND
            curr += timedelta(days=1)

        end_headers = ["æª”æ¬¡", "å®šåƒ¹", "å°ˆæ¡ˆåƒ¹"]; 
        for i, h in enumerate(end_headers):
            c_idx = end_c_start + i; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == total_cols: r = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r))

        date_start_col = 6
        for c_idx in range(date_start_col, total_cols + 1):
            c7 = ws.cell(header_start_row, c_idx); c7.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            if c_idx == date_start_col: set_border(c7, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c7, right=BS_MEDIUM)
            c8 = ws.cell(8, c_idx); c8.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            if c_idx == date_start_col: set_border(c8, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c8, right=BS_MEDIUM)

        curr_row = header_start_row + 2; grouped_data = {"å…¨å®¶å»£æ’­": sorted([r for r in rows if r["media"]=="å…¨å®¶å»£æ’­"], key=lambda x:x['seconds']), "æ–°é®®è¦–": sorted([r for r in rows if r["media"]=="æ–°é®®è¦–"], key=lambda x:x['seconds']), "å®¶æ¨‚ç¦": sorted([r for r in rows if r["media"]=="å®¶æ¨‚ç¦"], key=lambda x:x['seconds'])}
        total_store_count = 0; total_list_sum = 0
        for m_key, data in grouped_data.items():
            if not data: continue
            start_merge = curr_row; d_name = f"å…¨å®¶ä¾¿åˆ©å•†åº—\n{m_key}å»£å‘Š" if m_key != "å®¶æ¨‚ç¦" else "å®¶æ¨‚ç¦"
            for idx, r in enumerate(data):
                ws.row_dimensions[curr_row].height = 40; ws.cell(curr_row, 1, d_name).alignment = ALIGN_CENTER; ws.cell(curr_row, 2, r['region']).alignment = ALIGN_CENTER
                p_num = int(r.get('program_num', 0)); total_store_count += p_num; suffix = "é¢" if m_key == "æ–°é®®è¦–" else "åº—"; ws.cell(curr_row, 3, f"{p_num:,}{suffix}").alignment = ALIGN_CENTER
                ws.cell(curr_row, 4, r['daypart']).alignment = ALIGN_CENTER
                sec = r['seconds']; sec_txt = f"{sec}ç§’\nå½±ç‰‡/å½±åƒ 1920x1080 (mp4)" if m_key == "æ–°é®®è¦–" else f"{sec}ç§’å»£å‘Š"; c_spec = ws.cell(curr_row, 5, sec_txt); c_spec.alignment = ALIGN_CENTER; c_spec.font = Font(name=FONT_MAIN, size=10)
                row_sum = 0
                for d_idx in range(eff_days):
                    if d_idx < len(r['schedule']): val = r['schedule'][d_idx]; row_sum += val; c = ws.cell(curr_row, 6+d_idx); c.value = val; c.alignment = ALIGN_CENTER; c.font = FONT_STD; c.border = BORDER_ALL_THIN
                ws.cell(curr_row, end_c_start, row_sum).alignment = ALIGN_CENTER
                rate_val = r['rate_display']; 
                if isinstance(rate_val, (int, float)): total_list_sum += rate_val
                ws.cell(curr_row, end_c_start+1, rate_val).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER 
                pkg = r['pkg_display']; 
                if r.get('is_pkg_member'): pkg = r['nat_pkg_display'] if idx == 0 else None
                ws.cell(curr_row, end_c_start+2, pkg).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+2).alignment = ALIGN_CENTER
                for c_idx in range(1, total_cols + 1): c = ws.cell(curr_row, c_idx); c.font = FONT_STD; c.border = BORDER_ALL_THIN
                set_border(ws.cell(curr_row, 5), right=BS_MEDIUM); curr_row += 1
            ws.merge_cells(start_row=start_merge, start_column=1, end_row=curr_row-1, end_column=1)
            if data[0].get('is_pkg_member'): ws.merge_cells(start_row=start_merge, start_column=end_c_start+2, end_row=curr_row-1, end_column=end_c_start+2)
            draw_outer_border_fast(ws, start_merge, curr_row-1, 1, total_cols)

        ws.row_dimensions[curr_row].height = 40; ws.cell(curr_row, 3, total_store_count).number_format = FMT_NUMBER; ws.cell(curr_row, 3).alignment = ALIGN_CENTER; ws.cell(curr_row, 3).font = FONT_BOLD
        ws.cell(curr_row, 5, "Total").alignment = ALIGN_CENTER; ws.cell(curr_row, 5).font = FONT_BOLD
        for d_idx in range(eff_days): daily_sum = sum([r['schedule'][d_idx] for r in rows if d_idx < len(r['schedule'])]); c = ws.cell(curr_row, 6+d_idx); c.value = daily_sum; c.alignment = ALIGN_CENTER; c.font = FONT_BOLD
        ws.cell(curr_row, end_c_start, sum([sum(r['schedule']) for r in rows])).alignment = ALIGN_CENTER; ws.cell(curr_row, end_c_start).font = FONT_BOLD
        ws.cell(curr_row, end_c_start+1, total_list_sum).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).font = FONT_BOLD; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER
        ws.cell(curr_row, end_c_start+2, budget).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+2).font = FONT_BOLD; ws.cell(curr_row, end_c_start+2).alignment = ALIGN_CENTER
        for c_idx in range(1, total_cols+1): ws.cell(curr_row, c_idx).border = BORDER_ALL_THIN
        draw_outer_border_fast(ws, curr_row, curr_row, 1, total_cols)
        for c_idx in range(1, total_cols+1): set_border(ws.cell(curr_row, c_idx), bottom=BS_MEDIUM)
        set_border(ws.cell(curr_row, 5), right=BS_MEDIUM); curr_row += 1

        vat = int(budget * 0.05); grand_total = budget + vat
        footer_stack = [("è£½ä½œ", prod), ("5% VAT", vat), ("Grand Total", grand_total)]
        for lbl, val in footer_stack:
            ws.row_dimensions[curr_row].height = 30; c_l = ws.cell(curr_row, end_c_start+1); c_l.value = lbl; c_l.alignment = ALIGN_RIGHT; c_l.font = FONT_STD
            c_v = ws.cell(curr_row, end_c_start+2); c_v.value = val; c_v.number_format = FMT_MONEY; c_v.alignment = ALIGN_CENTER; c_v.font = FONT_BOLD 
            t, b, l, r = BS_THIN, BS_THIN, BS_MEDIUM, BS_THIN; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_l.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            t, b, l, r = BS_THIN, BS_THIN, BS_THIN, BS_MEDIUM; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_v.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            if lbl == "Grand Total":
                for c_idx in range(1, total_cols + 1): set_border(ws.cell(curr_row, c_idx), bottom=BS_MEDIUM)
            curr_row += 1
          
        curr_row += 1; start_footer = curr_row; r_col_start = 6 
        ws.row_dimensions[start_footer].height = 25; ws.cell(start_footer, r_col_start).value = "Remarksï¼š"
        ws.cell(start_footer, r_col_start).font = Font(name=FONT_MAIN, size=16, bold=True)
        r_row = start_footer
        for rm in remarks_list:
            r_row += 1; ws.row_dimensions[r_row].height = 25; is_red = rm.strip().startswith("1.") or rm.strip().startswith("4."); is_blue = rm.strip().startswith("6."); color = "000000"
            if is_red: color = "FF0000"
            if is_blue: color = "0000FF"
            c = ws.cell(r_row, r_col_start); c.value = rm; c.font = Font(name=FONT_MAIN, size=16, color=color)

        sig_col_start = 1
        ws.cell(start_footer, sig_col_start).value = "ä¹™        æ–¹ï¼š"; ws.cell(start_footer, sig_col_start).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+1, sig_col_start+1).value = client_name; ws.cell(start_footer+1, sig_col_start+1).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+2, sig_col_start).value = "çµ±ä¸€ç·¨è™Ÿï¼š"; ws.cell(start_footer+2, sig_col_start).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+2, sig_col_start+2).value = ""; ws.cell(start_footer+2, sig_col_start+2).font = Font(name=FONT_MAIN, size=16)
        ws.cell(start_footer+3, sig_col_start).value = "å®¢æˆ¶ç°½ç« ï¼š"; ws.cell(start_footer+3, sig_col_start).font = Font(name=FONT_MAIN, size=16)

        target_border_row = r_row + 2
        for c_idx in range(1, total_cols + 1): ws.cell(target_border_row, c_idx).border = Border(bottom=SIDE_DOUBLE)
        return target_border_row

    # Main Execution of Excel Generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
      
    # === ä¿®æ”¹é»ï¼šæ”¹ç”¨ä¸­æ–‡åˆ¤æ–· ===
    if format_type == "æ±å³":
        render_dongwu_optimized(ws, start_dt, end_dt, rows, final_budget_val, prod_cost)
    elif format_type == "è²æ´»":
        render_shenghuo_optimized(ws, start_dt, end_dt, rows, final_budget_val, prod_cost)
    else:
        render_bolin_optimized(ws, start_dt, end_dt, rows, final_budget_val, prod_cost)
    # ==========================

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# =========================================================
# 7. ä¸»ç¨‹å¼é‚è¼¯ (Main Execution Block)
# =========================================================

def main():
    try:
        with st.spinner("æ­£åœ¨è®€å– Google è©¦ç®—è¡¨è¨­å®šæª”..."):
            # === ä¿®æ”¹é»ï¼šè§£åŒ…æ–°å¢çš„ SALES_MAP ===
            STORE_COUNTS, STORE_COUNTS_NUM, PRICING_DB, SEC_FACTORS, SALES_MAP, err_msg = load_config_from_cloud(GSHEET_SHARE_URL)
          
        if err_msg:
            st.error(f"âŒ è¨­å®šæª”è¼‰å…¥å¤±æ•—: {err_msg}")
            st.stop()
          
        # --- Sidebar é‚è¼¯ (ç™»å…¥èˆ‡è¨­å®š) ---
        with st.sidebar:
            st.header("ğŸ•µï¸ ä¸»ç®¡ç™»å…¥")
            if not st.session_state.is_supervisor:
                pwd = st.text_input("è¼¸å…¥å¯†ç¢¼", type="password", key="pwd_input")
                if st.button("ç™»å…¥"):
                    if pwd == "1234":
                        st.session_state.is_supervisor = True
                        st.rerun()
                    else:
                        st.error("å¯†ç¢¼éŒ¯èª¤")
            else:
                st.success("âœ… ç›®å‰ç‹€æ…‹ï¼šä¸»ç®¡æ¨¡å¼")
                if st.button("ç™»å‡º"):
                    st.session_state.is_supervisor = False
                    st.rerun()
              
            st.markdown("---")
            st.subheader("â˜ï¸ Ragic é€£ç·šè¨­å®š")
              
            if st.session_state.is_supervisor:
                st.session_state.ragic_url = st.text_input("Ragic è¡¨å–®ç¶²å€", value=st.session_state.ragic_url)
                st.session_state.ragic_key = st.text_input("Ragic API Key", value=st.session_state.ragic_key, type="password")
            else:
                st.text_input("Ragic è¡¨å–®ç¶²å€", value=st.session_state.ragic_url, disabled=True)
              
            st.markdown("---")
            if st.button("ğŸ§¹ æ¸…é™¤å¿«å–"):
                st.cache_data.clear()
                st.rerun()

        # --- Main Content é‚è¼¯ (è¼¸å…¥èˆ‡å ±è¡¨) ---
        st.title("ğŸ“º åª’é«” Cue è¡¨ç”Ÿæˆå™¨ (v112.6 Sales Alias)")
        # === ä¿®æ”¹é»ï¼šé¡¯ç¤ºé¸é …æ”¹ç‚ºä¸­æ–‡ ===
        format_type = st.radio("é¸æ“‡æ ¼å¼", ["æ±å³", "è²æ´»", "é‰‘éœ–"], horizontal=True)
        # ==============================

        c1, c2, c3, c4, c5_sales = st.columns(5)
        with c1: client_name = st.text_input("å®¢æˆ¶åç¨±", "è¬åœ‹é€šè·¯")
        with c2: product_name = st.text_input("ç”¢å“åç¨±", "çµ±ä¸€å¸ƒä¸")
        with c3: total_budget_input = st.number_input("ç¸½é ç®— (æœªç¨… Net)", value=1000000, step=10000)
        with c4: prod_cost_input = st.number_input("è£½ä½œè²» (æœªç¨…)", value=0, step=1000)
        
        # === ä¿®æ”¹é»ï¼šæ¥­å‹™åç¨±æ”¹ç‚ºä¸‹æ‹‰é¸å–® ===
        with c5_sales: 
            # å–å¾— Sales Map çš„æ‰€æœ‰ Key (çœŸå) ä½œç‚ºé¸é …
            sales_options = list(SALES_MAP.keys()) if SALES_MAP else []
            sales_person = st.selectbox("æ¥­å‹™åç¨±", options=sales_options)
        # ================================

        # è™•ç†ä¸»ç®¡è¦†å¯«é ç®—åŠŸèƒ½
        final_budget_val = total_budget_input
        if st.session_state.is_supervisor:
            st.markdown("---")
            col_sup1, col_sup2 = st.columns([1, 2])
            with col_sup1: st.error("ğŸ”’ [ä¸»ç®¡] å°ˆæ¡ˆå„ªæƒ åƒ¹è¦†å¯«")
            with col_sup2:
                override_val = st.number_input("è¼¸å…¥æœ€çµ‚æˆäº¤åƒ¹", value=total_budget_input)
                if override_val != total_budget_input:
                    final_budget_val = override_val
                    st.caption(f"âš ï¸ ä½¿ç”¨ ${final_budget_val:,} çµç®—")
            st.markdown("---")

        c5, c6 = st.columns(2)
        with c5: start_date = st.date_input("é–‹å§‹æ—¥", datetime(2026, 1, 1))
        with c6: end_date = st.date_input("çµæŸæ—¥", datetime(2026, 1, 31))
        days_count = (end_date - start_date).days + 1
        st.info(f"ğŸ“… èµ°æœŸå…± **{days_count}** å¤©")

        with st.expander("ğŸ“ å‚™è¨»æ¬„ä½è¨­å®š", expanded=False):
            rc1, rc2, rc3 = st.columns(3)
            sign_deadline = rc1.date_input("å›ç°½æˆªæ­¢æ—¥", datetime.now() + timedelta(days=3))
            billing_month = rc2.text_input("è«‹æ¬¾æœˆä»½", "2026å¹´2æœˆ")
            payment_date = rc3.date_input("ä»˜æ¬¾å…Œç¾æ—¥", datetime(2026, 3, 31))

        st.markdown("### 3. åª’é«”æŠ•æ”¾è¨­å®š")
        col_cb1, col_cb2, col_cb3 = st.columns(3)
          
        def on_media_change():
            """åª’é«”å‹¾é¸è®Šæ›´æ™‚çš„è‡ªå‹•é…æ¯”é‚è¼¯"""
            active = []
            if st.session_state.get("cb_rad"): active.append("rad_share")
            if st.session_state.get("cb_fv"): active.append("fv_share")
            if st.session_state.get("cb_cf"): active.append("cf_share")
            if not active: return
            share = 100 // len(active)
            for key in active: st.session_state[key] = share
            rem = 100 - sum([st.session_state[k] for k in active])
            st.session_state[active[0]] += rem

        def on_slider_change(changed_key):
            """æ»‘æ¡¿æ‹‰å‹•æ™‚çš„è‡ªå‹•å¹³è¡¡é‚è¼¯ (åª’é«”ä½”æ¯”ç”¨)"""
            active = []
            if st.session_state.get("cb_rad"): active.append("rad_share")
            if st.session_state.get("cb_fv"): active.append("fv_share")
            if st.session_state.get("cb_cf"): active.append("cf_share")
            others = [k for k in active if k != changed_key]
            if not others:
                st.session_state[changed_key] = 100
            elif len(others) == 1:
                val = st.session_state[changed_key]
                st.session_state[others[0]] = max(0, 100 - val)
            elif len(others) == 2:
                val = st.session_state[changed_key]
                rem = max(0, 100 - val)
                k1, k2 = others[0], others[1]
                sum_others = st.session_state[k1] + st.session_state[k2]
                if sum_others == 0:
                    st.session_state[k1] = rem // 2
                    st.session_state[k2] = rem - st.session_state[k1]
                else:
                    ratio = st.session_state[k1] / sum_others
                    st.session_state[k1] = int(rem * ratio)
                    st.session_state[k2] = rem - st.session_state[k1]

        def on_sec_slider_change(media_prefix, changed_sec, all_secs):
            """æ–°å¢çš„ç§’æ•¸è‡ªå‹•å¹³è¡¡é‚è¼¯ã€‚"""
            key_changed = f"{media_prefix}{changed_sec}"
            new_val = st.session_state[key_changed]
            rem = 100 - new_val
              
            others = [s for s in all_secs if s != changed_sec]
            if not others:
                st.session_state[key_changed] = 100
                return

            current_sum_others = sum([st.session_state[f"{media_prefix}{s}"] for s in others])
              
            for i, s in enumerate(others):
                other_key = f"{media_prefix}{s}"
                if current_sum_others == 0:
                    new_other_val = rem // len(others)
                    if i == len(others) - 1:
                        new_other_val = rem - sum([st.session_state[f"{media_prefix}{x}"] for x in others if x != s])
                else:
                    ratio = st.session_state[other_key] / current_sum_others
                    new_other_val = int(rem * ratio)
                    if i == len(others) - 1:
                        allocated = new_val + sum([st.session_state[f"{media_prefix}{x}"] for x in others if x != s])
                        new_other_val = 100 - allocated
                  
                st.session_state[other_key] = max(0, new_other_val)

        is_rad = col_cb1.checkbox("å…¨å®¶å»£æ’­", key="cb_rad", on_change=on_media_change)
        is_fv = col_cb2.checkbox("æ–°é®®è¦–", key="cb_fv", on_change=on_media_change)
        is_cf = col_cb3.checkbox("å®¶æ¨‚ç¦", key="cb_cf", on_change=on_media_change)

        m1, m2, m3 = st.columns(3)
        config = {}
          
        # --- åª’é«”åƒæ•¸è¨­å®š UI å€å¡Š ---
        if is_rad:
            with m1:
                st.markdown("#### ğŸ“» å…¨å®¶å»£æ’­")
                is_nat = st.checkbox("å…¨çœè¯æ’­", True, key="rad_nat")
                regs = ["å…¨çœ"] if is_nat else st.multiselect("å€åŸŸ", REGIONS_ORDER, default=REGIONS_ORDER, key="rad_reg")
                if not is_nat and len(regs) == 6:
                    is_nat = True
                    regs = ["å…¨çœ"]
                    st.info("âœ… å·²é¸æ»¿6å€ï¼Œè‡ªå‹•è½‰ç‚ºå…¨çœè¯æ’­")
                  
                secs = st.multiselect("ç§’æ•¸", DURATIONS, [20], key="rad_sec")
                st.slider("é ç®— %", 0, 100, key="rad_share", on_change=on_slider_change, args=("rad_share",))
                  
                sorted_secs = sorted(secs)
                if sorted_secs:
                    keys_to_check = [f"rs_{s}" for s in sorted_secs]
                    if any(k not in st.session_state for k in keys_to_check):
                        default_val = 100 // len(sorted_secs)
                        for i, s in enumerate(sorted_secs):
                            k = f"rs_{s}"
                            if i == len(sorted_secs) - 1:
                                st.session_state[k] = 100 - (default_val * (len(sorted_secs)-1))
                            else:
                                st.session_state[k] = default_val
                      
                    sec_shares = {}
                    for s in sorted_secs:
                        st.slider(
                            f"{s}ç§’ %", 0, 100, 
                            key=f"rs_{s}", 
                            on_change=on_sec_slider_change, 
                            args=("rs_", s, sorted_secs)
                        )
                        sec_shares[s] = st.session_state[f"rs_{s}"]
                      
                    config["å…¨å®¶å»£æ’­"] = {"is_national": is_nat, "regions": regs, "sec_shares": sec_shares, "share": st.session_state.rad_share}

        if is_fv:
            with m2:
                st.markdown("#### ğŸ“º æ–°é®®è¦–")
                is_nat = st.checkbox("å…¨çœè¯æ’­", False, key="fv_nat")
                regs = ["å…¨çœ"] if is_nat else st.multiselect("å€åŸŸ", REGIONS_ORDER, default=["åŒ—å€"], key="fv_reg")
                if not is_nat and len(regs) == 6:
                    is_nat = True
                    regs = ["å…¨çœ"]
                    st.info("âœ… å·²é¸æ»¿6å€ï¼Œè‡ªå‹•è½‰ç‚ºå…¨çœè¯æ’­")
                  
                secs = st.multiselect("ç§’æ•¸", DURATIONS, [10], key="fv_sec")
                st.slider("é ç®— %", 0, 100, key="fv_share", on_change=on_slider_change, args=("fv_share",))
                  
                sorted_secs = sorted(secs)
                if sorted_secs:
                    keys_to_check = [f"fs_{s}" for s in sorted_secs]
                    if any(k not in st.session_state for k in keys_to_check):
                        default_val = 100 // len(sorted_secs)
                        for i, s in enumerate(sorted_secs):
                            k = f"fs_{s}"
                            if i == len(sorted_secs) - 1:
                                st.session_state[k] = 100 - (default_val * (len(sorted_secs)-1))
                            else:
                                st.session_state[k] = default_val
                      
                    sec_shares = {}
                    for s in sorted_secs:
                        st.slider(
                            f"{s}ç§’ %", 0, 100, 
                            key=f"fs_{s}", 
                            on_change=on_sec_slider_change, 
                            args=("fs_", s, sorted_secs)
                        )
                        sec_shares[s] = st.session_state[f"fs_{s}"]
                      
                    config["æ–°é®®è¦–"] = {"is_national": is_nat, "regions": regs, "sec_shares": sec_shares, "share": st.session_state.fv_share}

        if is_cf:
            with m3:
                st.markdown("#### ğŸ›’ å®¶æ¨‚ç¦")
                secs = st.multiselect("ç§’æ•¸", DURATIONS, [20], key="cf_sec")
                st.slider("é ç®— %", 0, 100, key="cf_share", on_change=on_slider_change, args=("cf_share",))
                  
                sorted_secs = sorted(secs)
                if sorted_secs:
                    keys_to_check = [f"cs_{s}" for s in sorted_secs]
                    if any(k not in st.session_state for k in keys_to_check):
                        default_val = 100 // len(sorted_secs)
                        for i, s in enumerate(sorted_secs):
                            k = f"cs_{s}"
                            if i == len(sorted_secs) - 1:
                                st.session_state[k] = 100 - (default_val * (len(sorted_secs)-1))
                            else:
                                st.session_state[k] = default_val
                      
                    sec_shares = {}
                    for s in sorted_secs:
                        st.slider(
                            f"{s}ç§’ %", 0, 100, 
                            key=f"cs_{s}", 
                            on_change=on_sec_slider_change, 
                            args=("cs_", s, sorted_secs)
                        )
                        sec_shares[s] = st.session_state[f"cs_{s}"]
                  
                    config["å®¶æ¨‚ç¦"] = {"regions": ["å…¨çœ"], "sec_shares": sec_shares, "share": st.session_state.cf_share}

        # --- é‹ç®—èˆ‡è¼¸å‡ºé‚è¼¯ ---
        if config:
            rows, total_list_accum, logs = calculate_plan_data(config, total_budget_input, days_count, PRICING_DB, SEC_FACTORS, STORE_COUNTS_NUM, REGIONS_ORDER)
            prod_cost = prod_cost_input 
            vat = int(round(final_budget_val * 0.05))
            grand_total = final_budget_val + vat
              
            p_str = f"{'ã€'.join([f'{s}ç§’' for s in sorted(list(set(r['seconds'] for r in rows)))])} {product_name}"
            rem = get_remarks_text(sign_deadline, billing_month, payment_date)
              
            html_preview = generate_html_preview(rows, days_count, start_date, end_date, client_name, p_str, format_type, rem, total_list_accum, grand_total, final_budget_val, prod_cost)
            st.components.v1.html(html_preview, height=700, scrolling=True)
              
            # ========== [æ–°å¢] æ’å…¥é‹ç®—é‚è¼¯é¢æ¿ ==========
            render_logic_panel(logs)
            # ===========================================
              
            st.markdown("---")
            st.subheader("ğŸ“¥ æª”æ¡ˆä¸‹è¼‰å€")
              
            xlsx_temp = generate_excel_from_scratch(format_type, start_date, end_date, client_name, product_name, rows, rem, final_budget_val, prod_cost, sales_person)
              
            col_dl1, col_dl2, col_ragic = st.columns([1, 1, 2])
              
            with col_dl2:
                pdf_bytes, method, err = xlsx_bytes_to_pdf_bytes(xlsx_temp)
                if pdf_bytes:
                    st.download_button(
                        f"ğŸ“¥ ä¸‹è¼‰ PDF", 
                        pdf_bytes, 
                        f"Cue_{safe_filename(client_name)}.pdf", 
                        key="pdf_dl_btn",
                        mime="application/pdf"
                    )
                else:
                    st.warning(f"PDF ç”Ÿæˆå¤±æ•—: {err}")

            with col_dl1:
                if st.session_state.is_supervisor:
                    st.download_button(
                        "ğŸ“¥ ä¸‹è¼‰ Excel (ä¸»ç®¡æ¬Šé™)", 
                        xlsx_temp, 
                        f"Cue_{safe_filename(client_name)}.xlsx", 
                        key="xlsx_dl_btn",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.info("ğŸ”’ Excel ä¸‹è¼‰åŠŸèƒ½åƒ…é™ä¸»ç®¡ä½¿ç”¨")

            with col_ragic:
                st.markdown("#### â˜ï¸ ä¸Šå‚³è‡³ Ragic")
                  
                if not st.session_state.ragic_confirm_state:
                    if st.button("ğŸš€ ä¸Šå‚³è³‡æ–™è‡³ Ragic", type="primary"):
                        st.session_state.ragic_confirm_state = True
                        st.rerun()
                else:
                    st.warning(f"å³å°‡ä¸Šå‚³ã€{client_name} - {product_name}ã€‘è‡³ Ragicï¼Œè«‹ç¢ºèªï¼Ÿ")
                    c_conf1, c_conf2 = st.columns(2)
                      
                    with c_conf1:
                        if st.button("âŒ å–æ¶ˆ"):
                            st.session_state.ragic_confirm_state = False
                            st.rerun()
                            
                    with c_conf2:
                        if st.button("âœ… ç¢ºèªä¸Šå‚³"):
                            with st.spinner("æ­£åœ¨ä¸Šå‚³è³‡æ–™èˆ‡æª”æ¡ˆ..."):
                                  
                                # Ragic æ¬„ä½å°ç…§è¡¨ (è«‹å‹¿éš¨æ„ä¿®æ”¹ ID)
                                RAGIC_MAP = {
                                    'client':     '1000080',  # å®¢æˆ¶åç¨±
                                    'product':    '1000081',  # ç”¢å“åç¨±
                                    'budget_raw': '1000082',  # ç¸½é ç®— (æœªç¨… Net)
                                    'budget_fin': '1000083',  # æœ€çµ‚æˆäº¤åƒ¹ (ä¸»ç®¡è¦†å¯«å¾Œ)
                                    'prod_cost':  '1000084',  # è£½ä½œè²»
                                    'format':     '1000078',  # å ±è¡¨æ ¼å¼ (Dongwu/Shenghuo/Bolin)
                                    'sales':      '1000079',  # æ¥­å‹™åç¨±
                                    'date_start': '1000085',  # é–‹å§‹æ—¥
                                    'date_end':   '1000086',  # çµæŸæ—¥
                                    'date_sign':  '1000087',  # å›ç°½æˆªæ­¢æ—¥
                                    'bill_month': '1000089',  # è«‹æ¬¾æœˆä»½
                                    'date_pay':   '1000088',  # ä»˜æ¬¾å…Œç¾æ—¥
                                    'details':    '1000090',  # è©³ç´°æŠ•æ”¾è¨­å®šæ‘˜è¦
                                    'file_xls':   '1000091',  # Excel æª”æ¡ˆä¸Šå‚³æ¬„ä½
                                    'file_pdf':   '1000092'   # PDF æª”æ¡ˆä¸Šå‚³æ¬„ä½
                                }

                                campaign_summary = format_campaign_details(config)
                                
                                # === ä¿®æ”¹é»ï¼šå–å¾—å°æ‡‰çš„ç¶½è™Ÿ (è‹¥ç„¡å‰‡ç”¨çœŸå) ===
                                sales_nickname = SALES_MAP.get(sales_person, sales_person)
                                # ========================================

                                data_payload = {
                                    RAGIC_MAP['client']:     client_name,
                                    RAGIC_MAP['product']:    product_name,
                                    RAGIC_MAP['budget_raw']: total_budget_input,
                                    RAGIC_MAP['budget_fin']: final_budget_val,
                                    RAGIC_MAP['prod_cost']:  prod_cost_input,
                                    RAGIC_MAP['format']:     format_type,
                                    RAGIC_MAP['sales']:      sales_nickname,  # é€™è£¡ä¸Šå‚³ç¶½è™Ÿ
                                    RAGIC_MAP['date_start']: str(start_date),
                                    RAGIC_MAP['date_end']:   str(end_date),
                                    RAGIC_MAP['date_sign']:  str(sign_deadline),
                                    RAGIC_MAP['bill_month']: billing_month,
                                    RAGIC_MAP['date_pay']:   str(payment_date),
                                    RAGIC_MAP['details']:    campaign_summary,
                                }

                                files_payload = {}
                                files_payload[RAGIC_MAP['file_xls']] = (
                                    f"Cue_{safe_filename(client_name)}.xlsx", 
                                    xlsx_temp, 
                                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                                  
                                if pdf_bytes:
                                    files_payload[RAGIC_MAP['file_pdf']] = (
                                        f"Cue_{safe_filename(client_name)}.pdf", 
                                        pdf_bytes, 
                                        'application/pdf'
                                    )

                                success, msg = upload_to_ragic(
                                    st.session_state.ragic_url,
                                    st.session_state.ragic_key,
                                    data_payload,
                                    files_payload
                                )
                                  
                                if success:
                                    st.success(msg)
                                    time.sleep(3)
                                else:
                                    st.error(f"ä¸Šå‚³å¤±æ•—: {msg}")
                            
                            st.session_state.ragic_confirm_state = False
                            time.sleep(1)
                            st.rerun()

    except Exception as e:
        st.error("ç¨‹å¼åŸ·è¡Œç™¼ç”ŸéŒ¯èª¤ï¼Œè«‹è¯çµ¡é–‹ç™¼è€…ã€‚")
        st.error(traceback.format_exc())

if __name__ == "__main__":
    main()
